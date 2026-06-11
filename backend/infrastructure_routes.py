"""Infrastructure module — internal-only.

Phase 1 endpoints:
    GET  /api/infrastructure/inboxes          — flat list of every email account
                                                visible to the requester (super_admin
                                                sees everything, infra-permitted users
                                                see their own — kept consistent with
                                                the rest of the platform).
    GET  /api/infrastructure/summary          — top-of-page cards: counts by status
                                                and total remaining capacity for
                                                today / this week / next 30 days.
    GET  /api/infrastructure/export           — Excel (.xlsx) or CSV export of the
                                                Inbox Inventory or the Domain
                                                Inventory.

Phase 2 (next iteration) will add per-account 30-day projection + a real
diversification-aware allocation engine. The schema in this file is built so
Phase 2 can drop straight into `_compute_inbox_status` and the daily-projection
helpers below.
"""

import csv
import io
from collections import defaultdict
from datetime import datetime, timedelta, timezone, date as _date_cls
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ---------- helpers ---------------------------------------------------------

def _domain_of(email: str) -> str:
    if not email or "@" not in email:
        return ""
    return email.rsplit("@", 1)[1].lower()


def _today_local_date(account: Dict[str, Any]) -> _date_cls:
    """RouteMail tracks daily counters per-account in UTC — we use UTC here so
    the numbers exactly match what the send worker enforces."""
    return datetime.now(timezone.utc).date()


def _emails_sent_today(account: Dict[str, Any]) -> int:
    """The send worker resets `daily_send_count` to 0 on the first send of each
    UTC day (see `last_send_date` check in server.py). If `last_send_date`
    isn't today, we treat the counter as 0 even though it's still on the doc.
    """
    last_str = account.get("last_send_date")
    today = _today_local_date(account).isoformat()
    if last_str and str(last_str)[:10] == today:
        return int(account.get("daily_send_count") or 0)
    return 0


def _compute_status(
    account: Dict[str, Any],
    sent_today: int,
    daily_limit: int,
    active_campaign_count: int,
) -> str:
    """Deterministic status mapping — see spec for the six categories.

    Phase 2 will fold in future-projection data; the function signature stays
    stable so swapping the input is cheap.
    """
    # Account-level overrides come first
    account_status = (account.get("status") or "").lower()
    if account_status == "disconnected" or account.get("last_error"):
        return "Risky"

    if account.get("warmup_enabled") and (account.get("warmup_status") or "").lower() in ("warming", "active"):
        # Warming inboxes are reported as Warming Up regardless of capacity
        # — they're not allocatable to cold campaigns.
        return "Warming Up"

    if account.get("paused") or account_status in ("paused", "paused_daily_limit"):
        return "Paused"

    remaining = max(daily_limit - sent_today, 0)
    if remaining <= 0:
        return "Fully Reserved"
    if sent_today > 0 or active_campaign_count > 0:
        # Some capacity already consumed today OR pre-committed to a campaign.
        # A finer "partially available" is the most useful real signal.
        return "Partially Available"
    return "Available"


# ---------- main inbox builder ---------------------------------------------

async def _load_inboxes(db, user_doc: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Returns Inbox rows the requester is allowed to see.

    Super admin → everything in the platform.
    Other infra-permitted users → their own accounts (same scoping as the rest
    of the app — no cross-tenant leakage).
    """
    is_admin = (user_doc.get("role") == "super_admin")
    q: Dict[str, Any] = {}
    if not is_admin:
        q["user_id"] = user_doc["user_id"]

    accounts = await db.email_accounts.find(q, {"_id": 0}).sort("email", 1).to_list(10000)

    # Pre-build active campaign assignments in one round trip rather than
    # N queries (matters once a workspace gets to a few hundred accounts).
    campaign_account_map: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

    camp_query = {"status": {"$in": ["running", "scheduled", "paused", "paused_daily_limit"]}}
    if not is_admin:
        camp_query["user_id"] = user_doc["user_id"]
    camps = await db.campaigns.find(
        camp_query,
        {"_id": 0, "campaign_id": 1, "name": 1, "status": 1, "account_ids": 1, "scheduled_at": 1},
    ).to_list(10000)
    for c in camps:
        for aid in (c.get("account_ids") or []):
            campaign_account_map[aid].append({
                "campaign_id": c.get("campaign_id"),
                "name": c.get("name"),
                "status": c.get("status"),
                "kind": "campaign",
                "scheduled_at": c.get("scheduled_at"),
            })

    drip_query = {"status": {"$in": ["running", "scheduled", "paused"]}}
    if not is_admin:
        drip_query["user_id"] = user_doc["user_id"]
    drips = await db.drip_campaigns.find(
        drip_query,
        {"_id": 0, "drip_id": 1, "name": 1, "status": 1, "account_ids": 1, "schedule": 1},
    ).to_list(10000)
    for d in drips:
        for aid in (d.get("account_ids") or []):
            campaign_account_map[aid].append({
                "campaign_id": d.get("drip_id"),
                "name": d.get("name"),
                "status": d.get("status"),
                "kind": "drip",
                "scheduled_at": (d.get("schedule") or {}).get("start_date"),
            })

    # Build the workspace-name map once (the platform doesn't have multi-
    # workspace today — display the owner email as the workspace label).
    user_ids = list({a.get("user_id") for a in accounts if a.get("user_id")})
    user_docs = await db.users.find(
        {"user_id": {"$in": user_ids}}, {"_id": 0, "user_id": 1, "email": 1, "name": 1}
    ).to_list(10000)
    workspace_by_uid = {u["user_id"]: (u.get("name") or u.get("email") or "—") for u in user_docs}

    rows: List[Dict[str, Any]] = []
    for acc in accounts:
        daily_limit = int(acc.get("daily_limit") or 50)
        sent_today = _emails_sent_today(acc)
        remaining = max(daily_limit - sent_today, 0)
        assignments = campaign_account_map.get(acc["account_id"], [])
        active_cnt = sum(1 for x in assignments if (x.get("status") or "").lower() in ("running", "scheduled"))
        status = _compute_status(acc, sent_today, daily_limit, active_cnt)

        warmup_status = "—"
        if acc.get("warmup_enabled"):
            warmup_status = (acc.get("warmup_status") or "active").title()

        rows.append({
            "account_id": acc["account_id"],
            "email": acc.get("email", ""),
            "domain": _domain_of(acc.get("email", "")),
            "ownership": acc.get("ownership") or "",
            "workspace": workspace_by_uid.get(acc.get("user_id"), "—"),
            "status": status,
            "daily_limit": daily_limit,
            "emails_sent_today": sent_today,
            "remaining_capacity": remaining,
            "active_campaign_count": active_cnt,
            "campaign_assignments": [
                {"name": x.get("name"), "kind": x.get("kind"), "status": x.get("status")}
                for x in assignments
            ],
            "warmup_status": warmup_status,
            "last_activity_at": acc.get("last_sent_at") or acc.get("imap_last_sync_at") or "",
            "user_id": acc.get("user_id"),
            "account_status": (acc.get("status") or "").lower(),
        })

    return rows


def _summarise(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Cards data: counts by status + capacity rollups."""
    inbox_counts = defaultdict(int)
    for r in rows:
        inbox_counts[r["status"]] += 1

    # Domain rollup → status of a domain is the WORST of its inbox statuses,
    # with a tie-break order. Same priority used in the export.
    domains: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in rows:
        if r["domain"]:
            domains[r["domain"]].append(r)

    PRIORITY = ["Fully Reserved", "Risky", "Paused", "Warming Up", "Partially Available", "Available"]

    def _domain_status(items: List[Dict[str, Any]]) -> str:
        seen = {x["status"] for x in items}
        for s in PRIORITY:
            if s in seen:
                # Special case — if *all* inboxes are Available, it's Available
                if s == "Available" and all(x["status"] == "Available" for x in items):
                    return "Available"
                # If *any* inbox has remaining capacity > 0, treat domain as
                # at least Partially Available rather than Fully Reserved.
                if s == "Fully Reserved" and any(x["remaining_capacity"] > 0 for x in items):
                    return "Partially Available"
                return s
        return "Available"

    domain_counts = defaultdict(int)
    domain_capacity: Dict[str, Dict[str, int]] = {}
    for dom, items in domains.items():
        s = _domain_status(items)
        domain_counts[s] += 1
        total = sum(x["daily_limit"] for x in items)
        used = sum(x["emails_sent_today"] for x in items)
        domain_capacity[dom] = {
            "total": total,
            "used": used,
            "remaining": max(total - used, 0),
            "inbox_count": len(items),
            "status": s,
        }

    remaining_today = sum(r["remaining_capacity"] for r in rows)
    # Phase 1 estimate — projecting forward is just (remaining_today × N).
    # Phase 2 replaces this with real per-day projection from drip_contacts.
    remaining_week = remaining_today * 7
    remaining_30 = remaining_today * 30

    return {
        "inbox_counts": {
            "available": inbox_counts.get("Available", 0),
            "partially_available": inbox_counts.get("Partially Available", 0),
            "fully_reserved": inbox_counts.get("Fully Reserved", 0),
            "warming_up": inbox_counts.get("Warming Up", 0),
            "paused": inbox_counts.get("Paused", 0),
            "risky": inbox_counts.get("Risky", 0),
            "total": len(rows),
        },
        "domain_counts": {
            "available": domain_counts.get("Available", 0),
            "partially_available": domain_counts.get("Partially Available", 0),
            "fully_reserved": domain_counts.get("Fully Reserved", 0),
            "total": len(domains),
        },
        "capacity": {
            "remaining_today": remaining_today,
            "remaining_week": remaining_week,
            "remaining_30_days": remaining_30,
            "note": "Week / 30-day capacity is a linear estimate; the Phase-2 "
                    "projection engine will replace this with per-day modelling "
                    "from drip_contacts.next_send_at.",
        },
        "domains": domain_capacity,
    }


# ---------- router ----------------------------------------------------------

def build_infrastructure_router(db, get_infra_user):
    """Phase 1 router. `get_infra_user` is the dependency that ensures the
    requester is a super_admin OR has `can_access_infrastructure=True`."""
    router = APIRouter(prefix="/infrastructure", tags=["infrastructure"])

    @router.get("/inboxes")
    async def list_inboxes(
        ownership: Optional[str] = Query(None),
        domain: Optional[str] = Query(None),
        status: Optional[str] = Query(None, description="Comma-separated allowed statuses"),
        warmup_status: Optional[str] = Query(None),
        min_remaining: Optional[int] = Query(None, description=">= N remaining capacity today"),
        search: Optional[str] = Query(None, description="Substring match on email"),
        user=Depends(get_infra_user),
    ):
        rows = await _load_inboxes(db, user)
        # apply filters in python — list is small enough (per-tenant) that
        # this is dramatically simpler than building dynamic Mongo queries.
        if ownership:
            o = ownership.strip().lower()
            rows = [r for r in rows if (r["ownership"] or "").lower() == o]
        if domain:
            d = domain.strip().lower()
            rows = [r for r in rows if r["domain"] == d]
        if status:
            allowed = {s.strip().lower() for s in status.split(",") if s.strip()}
            rows = [r for r in rows if r["status"].lower() in allowed]
        if warmup_status:
            w = warmup_status.strip().lower()
            rows = [r for r in rows if (r["warmup_status"] or "").lower() == w]
        if min_remaining is not None:
            rows = [r for r in rows if r["remaining_capacity"] >= min_remaining]
        if search:
            s = search.strip().lower()
            rows = [r for r in rows if s in (r["email"] or "").lower()]

        # Collect ownership + domain dropdowns from the unfiltered base set so
        # the UI can render filter selectors even when the current filter
        # zeros out the data.
        all_rows = await _load_inboxes(db, user)
        ownership_options = sorted({r["ownership"] for r in all_rows if r["ownership"]})
        domain_options = sorted({r["domain"] for r in all_rows if r["domain"]})

        return {
            "inboxes": rows,
            "filter_options": {
                "ownership": ownership_options,
                "domain": domain_options,
            },
            "total": len(rows),
        }

    @router.get("/summary")
    async def summary(user=Depends(get_infra_user)):
        rows = await _load_inboxes(db, user)
        return _summarise(rows)

    # -------- export --------

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill("solid", fgColor="4338CA")  # indigo-700

    def _xlsx_inbox_workbook(rows: List[Dict[str, Any]]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Inbox Inventory"
        headers = [
            "Email", "Domain", "Ownership", "Workspace", "Status",
            "Daily Limit", "Sent Today", "Remaining", "Active Campaigns",
            "Warmup Status", "Last Activity",
        ]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="left", vertical="center")
        for r_idx, row in enumerate(rows, 2):
            ws.cell(row=r_idx, column=1, value=row["email"])
            ws.cell(row=r_idx, column=2, value=row["domain"])
            ws.cell(row=r_idx, column=3, value=row["ownership"])
            ws.cell(row=r_idx, column=4, value=row["workspace"])
            ws.cell(row=r_idx, column=5, value=row["status"])
            ws.cell(row=r_idx, column=6, value=row["daily_limit"])
            ws.cell(row=r_idx, column=7, value=row["emails_sent_today"])
            ws.cell(row=r_idx, column=8, value=row["remaining_capacity"])
            ws.cell(row=r_idx, column=9, value=row["active_campaign_count"])
            ws.cell(row=r_idx, column=10, value=row["warmup_status"])
            ws.cell(row=r_idx, column=11, value=row["last_activity_at"])
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def _xlsx_domain_workbook(rows: List[Dict[str, Any]]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Domain Inventory"
        headers = ["Domain", "Inbox Count", "Total Daily Capacity", "Used Today", "Remaining Today", "Status"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
        summary = _summarise(rows)
        for r_idx, (dom, info) in enumerate(sorted(summary["domains"].items()), 2):
            ws.cell(row=r_idx, column=1, value=dom)
            ws.cell(row=r_idx, column=2, value=info["inbox_count"])
            ws.cell(row=r_idx, column=3, value=info["total"])
            ws.cell(row=r_idx, column=4, value=info["used"])
            ws.cell(row=r_idx, column=5, value=info["remaining"])
            ws.cell(row=r_idx, column=6, value=info["status"])
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def _csv_inbox(rows: List[Dict[str, Any]]) -> bytes:
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow([
            "Email", "Domain", "Ownership", "Workspace", "Status",
            "Daily Limit", "Sent Today", "Remaining", "Active Campaigns",
            "Warmup Status", "Last Activity",
        ])
        for r in rows:
            w.writerow([
                r["email"], r["domain"], r["ownership"], r["workspace"], r["status"],
                r["daily_limit"], r["emails_sent_today"], r["remaining_capacity"],
                r["active_campaign_count"], r["warmup_status"], r["last_activity_at"],
            ])
        return out.getvalue().encode()

    def _csv_domain(rows: List[Dict[str, Any]]) -> bytes:
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(["Domain", "Inbox Count", "Total Daily Capacity", "Used Today", "Remaining Today", "Status"])
        summary = _summarise(rows)
        for dom, info in sorted(summary["domains"].items()):
            w.writerow([dom, info["inbox_count"], info["total"], info["used"], info["remaining"], info["status"]])
        return out.getvalue().encode()

    @router.get("/export")
    async def export(
        type: str = Query("inboxes", description="inboxes | domains"),
        format: str = Query("xlsx", description="xlsx | csv"),
        user=Depends(get_infra_user),
    ):
        type = (type or "inboxes").lower()
        fmt = (format or "xlsx").lower()
        if type not in ("inboxes", "domains"):
            raise HTTPException(status_code=400, detail="type must be inboxes|domains")
        if fmt not in ("xlsx", "csv"):
            raise HTTPException(status_code=400, detail="format must be xlsx|csv")

        rows = await _load_inboxes(db, user)
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        ext = "xlsx" if fmt == "xlsx" else "csv"
        fname = f"RouteMail_Infrastructure_{type.title()}_{today}.{ext}"

        if type == "inboxes":
            content = _xlsx_inbox_workbook(rows) if fmt == "xlsx" else _csv_inbox(rows)
        else:
            content = _xlsx_domain_workbook(rows) if fmt == "xlsx" else _csv_domain(rows)

        media = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            if fmt == "xlsx" else "text/csv"
        )
        return StreamingResponse(
            iter([content]),
            media_type=media,
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    return router
