"""Iteration 50 — Drip schedule.start_date + Excel reporting (/api/reports/export).

Covers:
  • PUT /api/drip-campaigns/{id} accepts schedule.start_date and round-trips via GET.
  • /api/reports/export — XLSX content-type, 3 sheets (Summary, Campaigns, Drip Campaigns),
    correct headers per sheet, filename pattern, filters (from_date/to_date/campaign_type/status),
    invalid campaign_type → 400, drip status bucket counts vs drip_contacts.
  • Regression — /api/campaigns, /api/drip-campaigns, /api/auth/me,
    /api/subscription/prices, /api/admin/users all return 200 with the session.
"""
import io
import os
import uuid
from datetime import datetime, timedelta, timezone

import pytest
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from pymongo import MongoClient

load_dotenv("/app/backend/.env")

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
if not BASE_URL:
    with open("/app/frontend/.env") as f:
        for line in f:
            if line.startswith("REACT_APP_BACKEND_URL"):
                BASE_URL = line.split("=", 1)[1].strip().strip('"').rstrip("/")

MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]

SUPER_ADMIN_ID = "user_b3e333b0f467"  # dhruvmathur208@gmail.com
NORMAL_USER_ID = "user_35cc629e1385"  # drip.tester@example.com

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ---------------- fixtures -------------------------------------------------
@pytest.fixture(scope="module")
def mongo():
    c = MongoClient(MONGO_URL)
    yield c[DB_NAME]
    c.close()


def _make_session(db, user_id):
    token = f"TEST_iter50_{uuid.uuid4().hex}"
    db.user_sessions.insert_one({
        "session_token": token,
        "user_id": user_id,
        "created_at": datetime.now(timezone.utc),
        "expires_at": datetime.now(timezone.utc) + timedelta(hours=2),
    })
    return token


def _client(token):
    s = requests.Session()
    s.cookies.set("session_token", token)
    s.headers.update({"Authorization": f"Bearer {token}"})
    return s


@pytest.fixture(scope="module")
def admin_token(mongo):
    tok = _make_session(mongo, SUPER_ADMIN_ID)
    yield tok
    mongo.user_sessions.delete_one({"session_token": tok})


@pytest.fixture(scope="module")
def user_token(mongo):
    tok = _make_session(mongo, NORMAL_USER_ID)
    yield tok
    mongo.user_sessions.delete_one({"session_token": tok})


@pytest.fixture(scope="module")
def admin_client(admin_token):
    return _client(admin_token)


@pytest.fixture(scope="module")
def user_client(user_token):
    return _client(user_token)


# ---------------- regression: existing endpoints ---------------------------
class TestRegression:
    def test_auth_me(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/auth/me")
        assert r.status_code == 200, r.text
        assert r.json().get("user_id") == SUPER_ADMIN_ID

    def test_subscription_prices(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/subscription/prices")
        assert r.status_code == 200

    def test_admin_users(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/admin/users")
        assert r.status_code == 200

    def test_campaigns_list(self, user_client):
        r = user_client.get(f"{BASE_URL}/api/campaigns")
        assert r.status_code == 200
        assert isinstance(r.json(), list)

    def test_drip_campaigns_list(self, user_client):
        r = user_client.get(f"{BASE_URL}/api/drip-campaigns")
        assert r.status_code == 200
        assert isinstance(r.json(), list)


# ---------------- drip start_date round-trip -------------------------------
class TestDripStartDate:
    def _pick_drip(self, client):
        r = client.get(f"{BASE_URL}/api/drip-campaigns")
        assert r.status_code == 200
        drips = r.json()
        if not drips:
            pytest.skip("No drip campaigns exist for drip.tester")
        return drips[0]

    def test_put_persists_start_date(self, user_client, mongo):
        drip = self._pick_drip(user_client)
        drip_id = drip["drip_id"]
        prev_schedule = drip.get("schedule") or {}
        prev_start = prev_schedule.get("start_date")

        future = (datetime.now(timezone.utc) + timedelta(days=7)).strftime("%Y-%m-%d")
        new_schedule = dict(prev_schedule)
        new_schedule["start_date"] = future
        new_schedule.setdefault("timezone", "UTC")

        r = user_client.put(
            f"{BASE_URL}/api/drip-campaigns/{drip_id}",
            json={"schedule": new_schedule},
        )
        assert r.status_code == 200, r.text

        # GET to verify round-trip
        g = user_client.get(f"{BASE_URL}/api/drip-campaigns/{drip_id}")
        assert g.status_code == 200
        got_schedule = g.json().get("schedule") or {}
        assert got_schedule.get("start_date") == future, got_schedule

        # And verify Mongo persisted it (defence-in-depth)
        doc = mongo.drip_campaigns.find_one({"drip_id": drip_id}, {"_id": 0, "schedule": 1})
        assert (doc or {}).get("schedule", {}).get("start_date") == future

        # Restore previous value (None or original string) to keep state clean
        restore = dict(prev_schedule)
        if prev_start is None:
            restore.pop("start_date", None)
            restore["start_date"] = None
        else:
            restore["start_date"] = prev_start
        restore.setdefault("timezone", "UTC")
        user_client.put(
            f"{BASE_URL}/api/drip-campaigns/{drip_id}",
            json={"schedule": restore},
        )

    def test_start_date_accepts_string_no_strict_past_validation(self, user_client):
        drip = self._pick_drip(user_client)
        drip_id = drip["drip_id"]
        prev_schedule = drip.get("schedule") or {}
        past = "2020-01-15"
        new_schedule = dict(prev_schedule)
        new_schedule["start_date"] = past
        new_schedule.setdefault("timezone", "UTC")
        r = user_client.put(
            f"{BASE_URL}/api/drip-campaigns/{drip_id}",
            json={"schedule": new_schedule},
        )
        # Per spec: server stores whatever is given (frontend rejects past).
        assert r.status_code == 200, r.text
        g = user_client.get(f"{BASE_URL}/api/drip-campaigns/{drip_id}")
        assert (g.json().get("schedule") or {}).get("start_date") == past

        # restore
        restore = dict(prev_schedule)
        if "start_date" not in prev_schedule:
            restore["start_date"] = None
        restore.setdefault("timezone", "UTC")
        user_client.put(
            f"{BASE_URL}/api/drip-campaigns/{drip_id}",
            json={"schedule": restore},
        )


# ---------------- /api/reports/export --------------------------------------
class TestReportsExport:
    URL = None  # set in __init__-style class
    EXPECTED_CAMP_HEADERS = [
        "Campaign Name", "Campaign Type", "Status",
        "Created Date", "Scheduled Date", "Start Date",
        "Contacts Targeted", "Emails Sent",
        "Replies", "Bounce Count", "Unsubscribes",
        "Reply Rate", "List",
    ]
    EXPECTED_DRIP_HEADERS = [
        "Campaign Name", "Campaign Type", "Status",
        "Created Date", "Scheduled Start Date", "Start Date",
        "Total Steps",
        "Contacts Targeted", "Emails Sent",
        "Active Contacts", "Completed Contacts", "Stopped Contacts", "Currently Running",
        "Replies", "Bounce Count", "Unsubscribes",
        "Reply Rate", "List",
    ]

    def _get(self, client, **params):
        r = client.get(f"{BASE_URL}/api/reports/export", params=params)
        return r

    def test_returns_xlsx_with_three_sheets(self, user_client):
        r = self._get(user_client)
        assert r.status_code == 200, r.text[:500]
        ct = r.headers.get("content-type", "")
        assert XLSX_MIME in ct, ct
        # Open the workbook in memory
        wb = load_workbook(io.BytesIO(r.content))
        assert set(wb.sheetnames) == {"Summary", "Campaigns", "Drip Campaigns"}, wb.sheetnames

    def test_campaigns_sheet_headers(self, user_client):
        r = self._get(user_client)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb["Campaigns"]
        actual = [ws.cell(row=1, column=i + 1).value for i in range(len(self.EXPECTED_CAMP_HEADERS))]
        assert actual == self.EXPECTED_CAMP_HEADERS, actual

    def test_drip_sheet_headers(self, user_client):
        r = self._get(user_client)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb["Drip Campaigns"]
        actual = [ws.cell(row=1, column=i + 1).value for i in range(len(self.EXPECTED_DRIP_HEADERS))]
        assert actual == self.EXPECTED_DRIP_HEADERS, actual

    def test_content_disposition_filename(self, user_client):
        r = self._get(user_client, from_date="2025-01-01", to_date="2025-12-31")
        assert r.status_code == 200
        cd = r.headers.get("content-disposition", "")
        assert "RouteMail_Campaign_Report_2025-01-01_to_2025-12-31.xlsx" in cd, cd

    def test_invalid_campaign_type_returns_400(self, user_client):
        r = self._get(user_client, campaign_type="invalid")
        assert r.status_code == 400, r.text

    def test_campaign_type_campaigns_only(self, user_client):
        r = self._get(user_client, campaign_type="campaigns")
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws_drip = wb["Drip Campaigns"]
        # Headers row only -> max_row == 1 if no drip rows present
        assert ws_drip.max_row == 1, f"Expected only header row in Drip sheet, got max_row={ws_drip.max_row}"

    def test_campaign_type_drip_only(self, user_client):
        r = self._get(user_client, campaign_type="drip")
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws_camp = wb["Campaigns"]
        assert ws_camp.max_row == 1, f"Expected only header row in Campaigns sheet, got max_row={ws_camp.max_row}"

    def test_campaign_type_all_includes_both(self, user_client, mongo):
        r = self._get(user_client, campaign_type="all")
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        # If user has any drips, the Drip sheet should have rows
        drip_count_in_db = mongo.drip_campaigns.count_documents({"user_id": NORMAL_USER_ID})
        if drip_count_in_db > 0:
            assert wb["Drip Campaigns"].max_row >= 2, "Drip sheet should have rows in 'all' mode"

    def test_date_range_excludes_out_of_range(self, user_client):
        # Use a far-past narrow window; expect ~0 rows but workbook still valid
        r = self._get(user_client, from_date="1999-01-01", to_date="1999-01-02")
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        assert wb["Campaigns"].max_row == 1
        assert wb["Drip Campaigns"].max_row == 1

    def test_status_filter_comma_separated(self, user_client, mongo):
        # Pick a status that is unlikely to match → both sheets should have only header
        r = self._get(user_client, status="nonexistent_status_xyz")
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        assert wb["Campaigns"].max_row == 1
        assert wb["Drip Campaigns"].max_row == 1

        # And running,completed — should not error
        r2 = self._get(user_client, status="running,completed")
        assert r2.status_code == 200

    def test_drip_status_counts_match_drip_contacts(self, user_client, mongo):
        """For each drip row in the workbook, verify Active/Completed/Stopped/Replies/Bounce/Unsub
        match db.drip_contacts grouped by status."""
        r = self._get(user_client, campaign_type="drip")
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb["Drip Campaigns"]
        if ws.max_row < 2:
            pytest.skip("No drip rows to check counts against")
        # column indices (1-based): name=1, active=10, completed=11, stopped=12, replies=14, bounce=15, unsub=16
        # We need drip_id — but the sheet doesn't include it. Map by name.
        drips_in_db = list(mongo.drip_campaigns.find({"user_id": NORMAL_USER_ID}, {"_id": 0, "drip_id": 1, "name": 1}))
        name_to_id = {d["name"]: d["drip_id"] for d in drips_in_db}
        for row_idx in range(2, ws.max_row + 1):
            name = ws.cell(row=row_idx, column=1).value
            active = ws.cell(row=row_idx, column=10).value
            completed = ws.cell(row=row_idx, column=11).value
            stopped = ws.cell(row=row_idx, column=12).value
            replies = ws.cell(row=row_idx, column=14).value
            bounced = ws.cell(row=row_idx, column=15).value
            unsub = ws.cell(row=row_idx, column=16).value
            drip_id = name_to_id.get(name)
            if not drip_id:
                continue
            buckets = {"active": 0, "completed": 0, "replied": 0, "bounced": 0, "unsubscribed": 0}
            for c in mongo.drip_contacts.find({"drip_id": drip_id}, {"_id": 0, "status": 1}):
                st = (c.get("status") or "").lower()
                if st in buckets:
                    buckets[st] += 1
            assert active == buckets["active"], f"{name}: active {active}!={buckets['active']}"
            assert completed == buckets["completed"], f"{name}: completed {completed}!={buckets['completed']}"
            assert replies == buckets["replied"], f"{name}: replies {replies}!={buckets['replied']}"
            assert bounced == buckets["bounced"], f"{name}: bounced {bounced}!={buckets['bounced']}"
            assert unsub == buckets["unsubscribed"], f"{name}: unsub {unsub}!={buckets['unsubscribed']}"
            expected_stopped = buckets["replied"] + buckets["bounced"] + buckets["unsubscribed"]
            assert stopped == expected_stopped, f"{name}: stopped {stopped}!={expected_stopped}"

    def test_unauthenticated_blocked(self):
        # No cookie -> should be 401/403, definitely NOT 200 with xlsx
        r = requests.get(f"{BASE_URL}/api/reports/export")
        assert r.status_code in (401, 403), r.status_code
