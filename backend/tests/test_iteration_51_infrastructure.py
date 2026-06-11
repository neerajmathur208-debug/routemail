"""Iteration 51 — Phase 1 Infrastructure module backend tests.

Covers:
  • PUT /api/admin/users/{user_id}/infrastructure-permission (super_admin only)
  • /api/auth/me includes `can_access_infrastructure`
  • /api/admin/users response includes `can_access_infrastructure`
  • PUT /api/accounts/{account_id}/ownership (own-account OR super_admin)
  • GET /api/infrastructure/inboxes (auth gating + filters + per-tenant scoping)
  • GET /api/infrastructure/summary (counts + capacity + domains)
  • GET /api/infrastructure/export (xlsx + csv inboxes/domains, invalid → 400)
  • 403 on infra/* for non-permitted non-admin users
  • Regression: /campaigns, /drip-campaigns, /accounts, /auth/me,
    /subscription/prices all still 200.
"""
import io
import os
import uuid
from datetime import datetime, timedelta, timezone

import pytest
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from pymongo import MongoClient

load_dotenv("/app/backend/.env")

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
if not BASE_URL:
    with open("/app/frontend/.env") as f:
        for line in f:
            if line.startswith("REACT_APP_BACKEND_URL"):
                BASE_URL = line.split("=", 1)[1].strip().strip('"').rstrip("/")

MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]

SUPER_ADMIN_ID = "user_b3e333b0f467"   # dhruvmathur208@gmail.com
NORMAL_USER_ID = "user_35cc629e1385"   # drip.tester@example.com

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ---------------- fixtures -------------------------------------------------
@pytest.fixture(scope="module")
def mongo():
    c = MongoClient(MONGO_URL)
    yield c[DB_NAME]
    c.close()


def _make_session(db, user_id):
    token = f"TEST_iter51_{uuid.uuid4().hex}"
    db.user_sessions.insert_one({
        "session_token": token,
        "user_id": user_id,
        "created_at": datetime.now(timezone.utc),
        "expires_at": datetime.now(timezone.utc) + timedelta(hours=2),
    })
    return token


def _client(token):
    s = requests.Session()
    s.cookies.set("session_token", token)
    s.headers.update({"Authorization": f"Bearer {token}"})
    return s


@pytest.fixture(scope="module")
def admin_token(mongo):
    tok = _make_session(mongo, SUPER_ADMIN_ID)
    yield tok
    mongo.user_sessions.delete_one({"session_token": tok})


@pytest.fixture(scope="module")
def user_token(mongo):
    tok = _make_session(mongo, NORMAL_USER_ID)
    yield tok
    mongo.user_sessions.delete_one({"session_token": tok})


@pytest.fixture(scope="module")
def admin_client(admin_token):
    return _client(admin_token)


@pytest.fixture(scope="module")
def user_client(user_token):
    return _client(user_token)


@pytest.fixture(scope="module", autouse=True)
def _reset_normal_user_infra_flag(mongo):
    """Make sure drip.tester starts WITHOUT the infra flag, restore after."""
    original = mongo.users.find_one({"user_id": NORMAL_USER_ID}, {"can_access_infrastructure": 1})
    mongo.users.update_one(
        {"user_id": NORMAL_USER_ID},
        {"$set": {"can_access_infrastructure": False}},
    )
    yield
    if original and "can_access_infrastructure" in original:
        mongo.users.update_one(
            {"user_id": NORMAL_USER_ID},
            {"$set": {"can_access_infrastructure": original["can_access_infrastructure"]}},
        )
    else:
        mongo.users.update_one(
            {"user_id": NORMAL_USER_ID},
            {"$unset": {"can_access_infrastructure": ""}},
        )


# ============== /auth/me + /admin/users include flag ======================
class TestAuthMeFlag:
    def test_auth_me_super_admin_returns_flag(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/auth/me")
        assert r.status_code == 200, r.text
        data = r.json()
        assert "can_access_infrastructure" in data
        assert isinstance(data["can_access_infrastructure"], bool)

    def test_auth_me_normal_user_defaults_false(self, user_client):
        r = user_client.get(f"{BASE_URL}/api/auth/me")
        assert r.status_code == 200, r.text
        data = r.json()
        assert data.get("can_access_infrastructure") is False

    def test_admin_users_includes_flag(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/admin/users")
        assert r.status_code == 200, r.text
        users = r.json() if isinstance(r.json(), list) else r.json().get("users", [])
        assert len(users) > 0
        for u in users:
            assert "can_access_infrastructure" in u
            assert isinstance(u["can_access_infrastructure"], bool)


# ============== infrastructure-permission toggle ===========================
class TestInfraPermissionToggle:
    def test_non_admin_cannot_toggle(self, user_client):
        r = user_client.put(
            f"{BASE_URL}/api/admin/users/{NORMAL_USER_ID}/infrastructure-permission",
            json={"can_access_infrastructure": True},
        )
        assert r.status_code == 403, r.text

    def test_admin_can_toggle_on_and_off(self, admin_client, mongo):
        # ON
        r = admin_client.put(
            f"{BASE_URL}/api/admin/users/{NORMAL_USER_ID}/infrastructure-permission",
            json={"can_access_infrastructure": True},
        )
        assert r.status_code == 200, r.text
        body = r.json()
        assert body.get("can_access_infrastructure") is True

        doc = mongo.users.find_one({"user_id": NORMAL_USER_ID})
        assert doc.get("can_access_infrastructure") is True

        # OFF
        r2 = admin_client.put(
            f"{BASE_URL}/api/admin/users/{NORMAL_USER_ID}/infrastructure-permission",
            json={"can_access_infrastructure": False},
        )
        assert r2.status_code == 200
        assert r2.json().get("can_access_infrastructure") is False
        doc2 = mongo.users.find_one({"user_id": NORMAL_USER_ID})
        assert doc2.get("can_access_infrastructure") is False

    def test_missing_body_returns_400(self, admin_client):
        r = admin_client.put(
            f"{BASE_URL}/api/admin/users/{NORMAL_USER_ID}/infrastructure-permission",
            json={},
        )
        assert r.status_code == 400, r.text


# ============== ownership endpoint =========================================
class TestOwnership:
    def test_user_updates_own_account_ownership(self, user_client, mongo):
        acc = mongo.email_accounts.find_one({"user_id": NORMAL_USER_ID}, {"account_id": 1})
        assert acc, "Need at least one account on drip.tester"
        aid = acc["account_id"]

        r = user_client.put(
            f"{BASE_URL}/api/accounts/{aid}/ownership",
            json={"ownership": "TEST_iter51_label"},
        )
        assert r.status_code == 200, r.text
        assert r.json()["ownership"] == "TEST_iter51_label"
        doc = mongo.email_accounts.find_one({"account_id": aid}, {"ownership": 1})
        assert doc["ownership"] == "TEST_iter51_label"

        # clear
        r2 = user_client.put(
            f"{BASE_URL}/api/accounts/{aid}/ownership",
            json={"ownership": ""},
        )
        assert r2.status_code == 200
        assert r2.json()["ownership"] == ""

    def test_truncates_over_120_chars(self, user_client, mongo):
        acc = mongo.email_accounts.find_one({"user_id": NORMAL_USER_ID}, {"account_id": 1})
        aid = acc["account_id"]
        big = "x" * 200
        r = user_client.put(
            f"{BASE_URL}/api/accounts/{aid}/ownership",
            json={"ownership": big},
        )
        assert r.status_code == 200
        assert len(r.json()["ownership"]) == 120
        # cleanup
        user_client.put(f"{BASE_URL}/api/accounts/{aid}/ownership", json={"ownership": ""})

    def test_other_user_cannot_update_returns_404(self, user_client, mongo):
        # Find an account NOT owned by drip.tester
        other = mongo.email_accounts.find_one(
            {"user_id": {"$ne": NORMAL_USER_ID}}, {"account_id": 1}
        )
        if not other:
            pytest.skip("No cross-user account to test against")
        r = user_client.put(
            f"{BASE_URL}/api/accounts/{other['account_id']}/ownership",
            json={"ownership": "should not work"},
        )
        assert r.status_code == 404, r.text

    def test_super_admin_can_update_any_account(self, admin_client, mongo):
        acc = mongo.email_accounts.find_one({"user_id": NORMAL_USER_ID}, {"account_id": 1})
        aid = acc["account_id"]
        r = admin_client.put(
            f"{BASE_URL}/api/accounts/{aid}/ownership",
            json={"ownership": "TEST_admin_label"},
        )
        assert r.status_code == 200, r.text
        assert r.json()["ownership"] == "TEST_admin_label"
        # cleanup
        admin_client.put(f"{BASE_URL}/api/accounts/{aid}/ownership", json={"ownership": ""})


# ============== infra access gating ========================================
class TestInfraGating:
    def test_no_permission_user_blocked_all_infra(self, user_client, mongo):
        # ensure flag is False
        mongo.users.update_one(
            {"user_id": NORMAL_USER_ID},
            {"$set": {"can_access_infrastructure": False}},
        )
        for path in [
            "/api/infrastructure/inboxes",
            "/api/infrastructure/summary",
            "/api/infrastructure/export?type=inboxes&format=xlsx",
        ]:
            r = user_client.get(f"{BASE_URL}{path}")
            assert r.status_code == 403, f"{path} -> {r.status_code} {r.text}"

    def test_super_admin_can_access(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/inboxes")
        assert r.status_code == 200, r.text

    def test_infra_permitted_user_can_access(self, user_client, mongo):
        mongo.users.update_one(
            {"user_id": NORMAL_USER_ID},
            {"$set": {"can_access_infrastructure": True}},
        )
        try:
            r = user_client.get(f"{BASE_URL}/api/infrastructure/inboxes")
            assert r.status_code == 200, r.text
        finally:
            mongo.users.update_one(
                {"user_id": NORMAL_USER_ID},
                {"$set": {"can_access_infrastructure": False}},
            )


# ============== /infrastructure/inboxes ====================================
class TestInfraInboxes:
    def test_inbox_response_shape_and_fields(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/inboxes")
        assert r.status_code == 200
        data = r.json()
        assert "inboxes" in data and "filter_options" in data and "total" in data
        assert isinstance(data["inboxes"], list)
        assert data["total"] == len(data["inboxes"])
        if data["inboxes"]:
            row = data["inboxes"][0]
            for k in (
                "account_id", "email", "domain", "ownership", "workspace",
                "status", "daily_limit", "emails_sent_today",
                "remaining_capacity", "active_campaign_count",
                "campaign_assignments", "warmup_status", "last_activity_at",
            ):
                assert k in row, f"Missing key {k}"
        assert "ownership" in data["filter_options"]
        assert "domain" in data["filter_options"]

    def test_status_categories_present(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/inboxes")
        data = r.json()
        statuses = {row["status"] for row in data["inboxes"]}
        # spec says at minimum Available + Warming Up
        valid = {"Available", "Warming Up", "Partially Available",
                 "Fully Reserved", "Paused", "Risky"}
        assert statuses.issubset(valid), f"Unexpected statuses: {statuses - valid}"
        assert "Available" in statuses, f"Expected Available; got {statuses}"

    def test_filter_status_available(self, admin_client):
        r = admin_client.get(
            f"{BASE_URL}/api/infrastructure/inboxes?status=Available"
        )
        assert r.status_code == 200
        data = r.json()
        for row in data["inboxes"]:
            assert row["status"] == "Available"

    def test_filter_min_remaining(self, admin_client):
        r = admin_client.get(
            f"{BASE_URL}/api/infrastructure/inboxes?min_remaining=1"
        )
        assert r.status_code == 200
        for row in r.json()["inboxes"]:
            assert row["remaining_capacity"] >= 1

    def test_filter_domain(self, admin_client):
        base = admin_client.get(f"{BASE_URL}/api/infrastructure/inboxes").json()
        domains = base["filter_options"]["domain"]
        if not domains:
            pytest.skip("No domains in dataset")
        d = domains[0]
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/inboxes?domain={d}")
        assert r.status_code == 200
        for row in r.json()["inboxes"]:
            assert row["domain"] == d

    def test_filter_search_email_substring(self, admin_client):
        base = admin_client.get(f"{BASE_URL}/api/infrastructure/inboxes").json()
        if not base["inboxes"]:
            pytest.skip()
        email = base["inboxes"][0]["email"]
        sub = email.split("@")[0][:4]
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/inboxes?search={sub}")
        assert r.status_code == 200
        for row in r.json()["inboxes"]:
            assert sub.lower() in row["email"].lower()

    def test_non_admin_sees_only_own_accounts(self, user_client, mongo):
        mongo.users.update_one(
            {"user_id": NORMAL_USER_ID},
            {"$set": {"can_access_infrastructure": True}},
        )
        try:
            r = user_client.get(f"{BASE_URL}/api/infrastructure/inboxes")
            assert r.status_code == 200
            for row in r.json()["inboxes"]:
                # Workspace label is the owner's name/email — but the user_id
                # field on the row should match
                # The row has no user_id field exposed; verify via account_id
                # belongs to this user in Mongo
                acc = mongo.email_accounts.find_one(
                    {"account_id": row["account_id"]}, {"user_id": 1}
                )
                assert acc["user_id"] == NORMAL_USER_ID, (
                    f"Cross-tenant leak: account {row['account_id']} belongs "
                    f"to {acc['user_id']}"
                )
        finally:
            mongo.users.update_one(
                {"user_id": NORMAL_USER_ID},
                {"$set": {"can_access_infrastructure": False}},
            )


# ============== /infrastructure/summary ====================================
class TestInfraSummary:
    def test_summary_shape(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/summary")
        assert r.status_code == 200
        d = r.json()
        for k in ("inbox_counts", "domain_counts", "capacity", "domains"):
            assert k in d

        ic = d["inbox_counts"]
        for k in ("available", "partially_available", "fully_reserved",
                  "warming_up", "paused", "risky", "total"):
            assert k in ic
        dc = d["domain_counts"]
        for k in ("available", "partially_available", "fully_reserved", "total"):
            assert k in dc
        cap = d["capacity"]
        for k in ("remaining_today", "remaining_week", "remaining_30_days"):
            assert k in cap and isinstance(cap[k], int)
        # Phase-1 documented linear estimate
        assert cap["remaining_week"] == cap["remaining_today"] * 7
        assert cap["remaining_30_days"] == cap["remaining_today"] * 30

        # domains keyed by domain name
        for dom, info in d["domains"].items():
            for k in ("total", "used", "remaining", "inbox_count", "status"):
                assert k in info

    def test_summary_total_matches_inbox_list(self, admin_client):
        s = admin_client.get(f"{BASE_URL}/api/infrastructure/summary").json()
        ib = admin_client.get(f"{BASE_URL}/api/infrastructure/inboxes").json()
        assert s["inbox_counts"]["total"] == ib["total"]


# ============== /infrastructure/export =====================================
class TestInfraExport:
    def test_inboxes_xlsx(self, admin_client):
        r = admin_client.get(
            f"{BASE_URL}/api/infrastructure/export?type=inboxes&format=xlsx"
        )
        assert r.status_code == 200
        assert XLSX_MIME in r.headers.get("content-type", "")
        cd = r.headers.get("content-disposition", "")
        assert "RouteMail_Infrastructure_Inboxes_" in cd
        assert cd.endswith('.xlsx"')

        wb = load_workbook(io.BytesIO(r.content))
        assert "Inbox Inventory" in wb.sheetnames
        ws = wb["Inbox Inventory"]
        headers = [c.value for c in ws[1]]
        assert headers == [
            "Email", "Domain", "Ownership", "Workspace", "Status",
            "Daily Limit", "Sent Today", "Remaining", "Active Campaigns",
            "Warmup Status", "Last Activity",
        ]
        ib = admin_client.get(f"{BASE_URL}/api/infrastructure/inboxes").json()
        # data rows = total inboxes (ws.max_row includes header)
        assert ws.max_row - 1 == ib["total"]

    def test_domains_csv(self, admin_client):
        r = admin_client.get(
            f"{BASE_URL}/api/infrastructure/export?type=domains&format=csv"
        )
        assert r.status_code == 200
        assert "text/csv" in r.headers.get("content-type", "")
        first_line = r.text.splitlines()[0].strip()
        assert first_line == "Domain,Inbox Count,Total Daily Capacity,Used Today,Remaining Today,Status"

    def test_invalid_type(self, admin_client):
        r = admin_client.get(
            f"{BASE_URL}/api/infrastructure/export?type=bogus&format=xlsx"
        )
        assert r.status_code == 400

    def test_invalid_format(self, admin_client):
        r = admin_client.get(
            f"{BASE_URL}/api/infrastructure/export?type=inboxes&format=bogus"
        )
        assert r.status_code == 400


# ============== regression =================================================
class TestRegression:
    @pytest.mark.parametrize("path", [
        "/api/auth/me",
        "/api/campaigns",
        "/api/drip-campaigns",
        "/api/accounts",
        "/api/subscription/prices",
    ])
    def test_endpoint_still_200(self, admin_client, path):
        r = admin_client.get(f"{BASE_URL}{path}")
        assert r.status_code == 200, f"{path} -> {r.status_code} {r.text[:200]}"
