"""Iteration 55 — Batch-Based Weekly Sending Planner backend tests.

Covers POST /api/infrastructure/planner/batch and /planner/batch/export.
"""
import io
import os
import uuid
from datetime import datetime, timedelta, timezone

import pytest
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from pymongo import MongoClient

load_dotenv("/app/backend/.env")

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
if not BASE_URL:
    with open("/app/frontend/.env") as f:
        for line in f:
            if line.startswith("REACT_APP_BACKEND_URL"):
                BASE_URL = line.split("=", 1)[1].strip().strip('"').rstrip("/")

MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]

SUPER_ADMIN_ID = "user_b3e333b0f467"
NORMAL_USER_ID = "user_35cc629e1385"

BATCH_URL = f"{BASE_URL}/api/infrastructure/planner/batch"
EXPORT_URL = f"{BASE_URL}/api/infrastructure/planner/batch/export"


@pytest.fixture(scope="module")
def mongo():
    c = MongoClient(MONGO_URL)
    yield c[DB_NAME]
    c.close()


def _make_session(db, user_id):
    token = f"TEST_iter55_{uuid.uuid4().hex}"
    db.user_sessions.insert_one({
        "session_token": token,
        "user_id": user_id,
        "created_at": datetime.now(timezone.utc),
        "expires_at": datetime.now(timezone.utc) + timedelta(hours=2),
    })
    return token


def _client(token):
    s = requests.Session()
    s.cookies.set("session_token", token)
    s.headers.update({"Authorization": f"Bearer {token}"})
    return s


@pytest.fixture(scope="module")
def admin_token(mongo):
    tok = _make_session(mongo, SUPER_ADMIN_ID)
    yield tok
    mongo.user_sessions.delete_one({"session_token": tok})


@pytest.fixture(scope="module")
def user_token(mongo):
    tok = _make_session(mongo, NORMAL_USER_ID)
    mongo.users.update_one(
        {"user_id": NORMAL_USER_ID},
        {"$set": {"can_access_infrastructure": False}},
    )
    yield tok
    mongo.user_sessions.delete_one({"session_token": tok})


@pytest.fixture(scope="module")
def admin_client(admin_token):
    return _client(admin_token)


@pytest.fixture(scope="module")
def user_client(user_token):
    return _client(user_token)


SPEC_PAYLOAD = {
    "leads": 4000, "steps": 3, "delay_days": 7,
    "sending_days": [0, 1, 2, 3, 4], "start_date": "2026-09-14",
    "accounts": 20, "daily_limit_per_account": 40,
}


# --- spec exact example ---
class TestSpecExample:
    def test_summary_and_batches(self, admin_client):
        r = admin_client.post(BATCH_URL, json=SPEC_PAYLOAD)
        assert r.status_code == 200, r.text
        d = r.json()
        s = d["summary"]
        assert s["total_batches"] == 5
        assert s["daily_capacity"] == 800
        assert s["total_emails"] == 12000
        assert s["status"] == "Ready"
        assert s["total_leads"] == 4000
        # 5 batches of 800
        assert len(d["batches"]) == 5
        assert all(b["leads"] == 800 for b in d["batches"])
        # Step-1 dates Mon..Fri Sep 14..18
        expected = ["2026-09-14", "2026-09-15", "2026-09-16", "2026-09-17", "2026-09-18"]
        assert [b["step_1_date"] for b in d["batches"]] == expected

    def test_schedule_15_rows_all_ready(self, admin_client):
        r = admin_client.post(BATCH_URL, json=SPEC_PAYLOAD)
        d = r.json()
        sched = d["schedule"]
        assert len(sched) == 15
        assert all(row["status"] == "Ready" for row in sched)

    def test_followup_alignment(self, admin_client):
        r = admin_client.post(BATCH_URL, json=SPEC_PAYLOAD)
        d = r.json()
        # Build (batch, step) -> date map
        m = {(row["batch"], row["step"]): row["date"] for row in d["schedule"]}
        # Batch 1 step 2 = 2026-09-21 (Mon), step 3 = 2026-09-28 (Mon)
        assert m[(1, 2)] == "2026-09-21"
        assert m[(1, 3)] == "2026-09-28"
        # Batch 5 step 3 = 2026-10-02 (Fri)
        assert m[(5, 3)] == "2026-10-02"


# --- uneven divide ---
class TestUnevenDivide:
    def test_4250_yields_6_batches(self, admin_client):
        payload = {
            "leads": 4250, "steps": 1, "delay_days": 7,
            "sending_days": [0, 1, 2, 3, 4], "start_date": "2026-09-14",
            "accounts": 20, "daily_limit_per_account": 40,
        }
        r = admin_client.post(BATCH_URL, json=payload)
        assert r.status_code == 200, r.text
        d = r.json()
        sizes = [b["leads"] for b in d["batches"]]
        assert sizes == [800, 800, 800, 800, 800, 250]
        assert d["summary"]["total_batches"] == 6


# --- weekday snapping ---
class TestWeekdaySnap:
    def test_step2_snaps_forward_to_mon(self, admin_client):
        payload = {
            "leads": 50, "steps": 2, "delay_days": 5,
            "sending_days": [0, 1, 2, 3, 4], "start_date": "2026-09-14",
            "accounts": 1, "daily_limit_per_account": 100,
        }
        r = admin_client.post(BATCH_URL, json=payload)
        assert r.status_code == 200, r.text
        m = {(row["batch"], row["step"]): row["date"] for row in r.json()["schedule"]}
        assert m[(1, 1)] == "2026-09-14"
        assert m[(1, 2)] == "2026-09-21"

    def test_step2_stays_saturday_when_weekends_allowed(self, admin_client):
        payload = {
            "leads": 50, "steps": 2, "delay_days": 5,
            "sending_days": [0, 1, 2, 3, 4, 5, 6], "start_date": "2026-09-14",
            "accounts": 1, "daily_limit_per_account": 100,
        }
        r = admin_client.post(BATCH_URL, json=payload)
        assert r.status_code == 200, r.text
        m = {(row["batch"], row["step"]): row["date"] for row in r.json()["schedule"]}
        assert m[(1, 2)] == "2026-09-19"  # Sat Sep 19


# --- real inbox pool ---
class TestRealPool:
    def test_real_pool_uses_inbox_daily_limits(self, admin_client, mongo):
        # Pick first 2 visible accounts for the super admin
        # Use the /inboxes endpoint so we get the loader-computed daily_limit
        # (matches what _batch_plan uses internally — raw email_accounts.daily_limit
        # is not the field the infra loader exposes).
        inb = admin_client.get(f"{BASE_URL}/api/infrastructure/inboxes").json()
        rows = inb if isinstance(inb, list) else (inb.get("inboxes") or [])
        assert len(rows) >= 2, "need at least 2 inboxes for super admin"
        picked = rows[:2]
        ids = [a["account_id"] for a in picked]
        expected_cap = sum(int(a.get("daily_limit", 0)) for a in picked)
        payload = {
            "leads": 100, "steps": 1, "delay_days": 7,
            "sending_days": [0, 1, 2, 3, 4], "start_date": "2026-09-14",
            "account_ids": ids,
        }
        r = admin_client.post(BATCH_URL, json=payload)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["inputs"]["pool_source"] == "real_inbox_pool"
        assert d["summary"]["daily_capacity"] == expected_cap

    def test_unknown_account_ids_400(self, admin_client):
        payload = {
            "leads": 100, "steps": 1, "delay_days": 7,
            "sending_days": [0, 1, 2, 3, 4], "start_date": "2026-09-14",
            "account_ids": ["acc_unknown_zzz1", "acc_unknown_zzz2"],
        }
        r = admin_client.post(BATCH_URL, json=payload)
        assert r.status_code == 400


# --- validation ---
class TestValidation:
    def test_leads_zero_422(self, admin_client):
        payload = dict(SPEC_PAYLOAD, leads=0)
        r = admin_client.post(BATCH_URL, json=payload)
        assert r.status_code == 422

    def test_sending_days_empty_is_allowed_and_falls_back(self, admin_client):
        payload = dict(SPEC_PAYLOAD, sending_days=[])
        r = admin_client.post(BATCH_URL, json=payload)
        # Pydantic accepts [] (no min_length), planner falls back to default
        assert r.status_code == 200, r.text

    def test_bad_start_date_400(self, admin_client):
        payload = dict(SPEC_PAYLOAD, start_date="not-a-date")
        r = admin_client.post(BATCH_URL, json=payload)
        assert r.status_code == 400

    def test_missing_both_pool_options_400(self, admin_client):
        payload = {
            "leads": 100, "steps": 1, "delay_days": 7,
            "sending_days": [0, 1, 2, 3, 4], "start_date": "2026-09-14",
        }
        r = admin_client.post(BATCH_URL, json=payload)
        assert r.status_code == 400
        detail = (r.json().get("detail") or "").lower()
        assert "account_ids" in detail and ("accounts" in detail and "daily_limit_per_account" in detail)


# --- capacity conflict detection ---
class TestCapacityConflict:
    def test_overlap_detection(self, admin_client):
        payload = {
            "leads": 1600, "steps": 2, "delay_days": 1,
            "sending_days": [0, 1, 2, 3, 4], "start_date": "2026-09-14",
            "accounts": 20, "daily_limit_per_account": 40,
        }
        r = admin_client.post(BATCH_URL, json=payload)
        assert r.status_code == 200, r.text
        d = r.json()
        # status should be downgraded
        assert d["summary"]["status"] in {"Partial Capacity", "Insufficient Capacity"}
        # at least one warning mentions all 4 key strings
        joined = " | ".join(d["warnings"])
        assert "Capacity exceeded" in joined
        assert "Required" in joined and "Available" in joined and "Shortfall" in joined
        # Some row must be tagged non-Ready on Tue (2026-09-15)
        bad_rows = [r2 for r2 in d["schedule"] if r2["status"] != "Ready"]
        assert any(r2["date"] == "2026-09-15" for r2 in bad_rows)
        # 1600 required vs 800 available on the colliding date
        tue = [r2 for r2 in d["schedule"] if r2["date"] == "2026-09-15"][0]
        assert tue["required_capacity"] == 1600
        assert tue["available_capacity"] == 800


# --- exports ---
class TestExports:
    def test_xlsx_export(self, admin_client):
        r = admin_client.post(f"{EXPORT_URL}?format=xlsx", json=SPEC_PAYLOAD)
        assert r.status_code == 200, r.text
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml" in ct
        wb = load_workbook(io.BytesIO(r.content))
        assert wb.sheetnames[0] == "Summary"
        assert "Schedule" in wb.sheetnames
        # Summary first
        assert wb.sheetnames.index("Summary") < wb.sheetnames.index("Schedule")
        ws = wb["Schedule"]
        # Header + 15 data rows
        assert ws.max_row == 16
        headers = [c.value for c in ws[1]]
        assert headers == [
            "Date", "Day", "Batch", "Step", "Leads Scheduled",
            "Required Capacity", "Available Capacity", "Shortfall", "Status",
        ]

    def test_csv_export(self, admin_client):
        r = admin_client.post(f"{EXPORT_URL}?format=csv", json=SPEC_PAYLOAD)
        assert r.status_code == 200, r.text
        assert "text/csv" in r.headers.get("content-type", "")
        # header + 15 schedule rows = 16 lines (trailing newline tolerated)
        lines = [ln for ln in r.content.decode().splitlines() if ln.strip()]
        assert len(lines) == 16

    def test_invalid_export_format_400(self, admin_client):
        r = admin_client.post(f"{EXPORT_URL}?format=pdf", json=SPEC_PAYLOAD)
        assert r.status_code == 400


# --- permission enforcement ---
class TestPermissions:
    def test_non_permitted_user_403_on_batch(self, user_client):
        r = user_client.post(BATCH_URL, json=SPEC_PAYLOAD)
        assert r.status_code == 403

    def test_non_permitted_user_403_on_export(self, user_client):
        r = user_client.post(f"{EXPORT_URL}?format=csv", json=SPEC_PAYLOAD)
        assert r.status_code == 403


# --- regression: standard planner still works ---
class TestRegression:
    def test_standard_planner_ready(self, admin_client):
        r = admin_client.post(f"{BASE_URL}/api/infrastructure/planner", json={
            "leads": 1000, "steps": 3, "duration_days": 60,
        })
        assert r.status_code == 200, r.text
        assert r.json()["status"] == "Ready"

    def test_phase1_summary_still_works(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/summary")
        assert r.status_code == 200, r.text

    def test_phase2_inboxes_still_works(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/inboxes")
        assert r.status_code == 200, r.text
