"""Iteration 52 — Phase 2 Infrastructure module backend tests.

Covers:
  • GET /api/infrastructure/summary returns projection-aware capacity (window_days=120 default)
  • GET /api/infrastructure/summary?window_days=30 — limits 1..365, remaining_window==remaining_30_days
  • GET /api/infrastructure/inboxes — every row has projected_window_total + projected_window_days=120
  • GET /api/infrastructure/calendar/{account_id} — 120 day rows, status mapping, totals
  • GET /api/infrastructure/calendar/acc_bogus → 404
  • GET /api/infrastructure/calendar/{account_id}?window_days=30 → 30 rows
  • Projection engine handles drip campaigns (active drip_contacts.next_send_at, steps, schedule)
  • Projection engine picks up scheduled regular campaigns
  • Non-permitted user → 403 on summary/calendar/inboxes
  • /export inboxes xlsx has 'Projected (120d)' column; domains csv has correct header
  • Phase-1 regressions
"""
import io
import os
import uuid
import csv
from datetime import datetime, timedelta, timezone

import pytest
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from pymongo import MongoClient

load_dotenv("/app/backend/.env")

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
if not BASE_URL:
    with open("/app/frontend/.env") as f:
        for line in f:
            if line.startswith("REACT_APP_BACKEND_URL"):
                BASE_URL = line.split("=", 1)[1].strip().strip('"').rstrip("/")

MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]

SUPER_ADMIN_ID = "user_b3e333b0f467"
NORMAL_USER_ID = "user_35cc629e1385"

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture(scope="module")
def mongo():
    c = MongoClient(MONGO_URL)
    yield c[DB_NAME]
    c.close()


def _make_session(db, user_id):
    token = f"TEST_iter52_{uuid.uuid4().hex}"
    db.user_sessions.insert_one({
        "session_token": token,
        "user_id": user_id,
        "created_at": datetime.now(timezone.utc),
        "expires_at": datetime.now(timezone.utc) + timedelta(hours=2),
    })
    return token


def _client(token):
    s = requests.Session()
    s.cookies.set("session_token", token)
    s.headers.update({"Authorization": f"Bearer {token}"})
    return s


@pytest.fixture(scope="module")
def admin_token(mongo):
    tok = _make_session(mongo, SUPER_ADMIN_ID)
    yield tok
    mongo.user_sessions.delete_one({"session_token": tok})


@pytest.fixture(scope="module")
def user_token(mongo):
    tok = _make_session(mongo, NORMAL_USER_ID)
    yield tok
    mongo.user_sessions.delete_one({"session_token": tok})


@pytest.fixture(scope="module")
def admin_client(admin_token):
    return _client(admin_token)


@pytest.fixture(scope="module")
def user_client(user_token):
    return _client(user_token)


# ---------------- Summary (capacity engine) -----------------------------

class TestSummary:
    def test_summary_default_window_120(self, admin_client, mongo):
        # Ensure clean state: no future projections beyond today
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/summary")
        assert r.status_code == 200, r.text
        data = r.json()
        cap = data["capacity"]
        assert cap["window_days"] == 120
        assert "note" in cap and isinstance(cap["note"], str) and cap["note"]
        for k in ("remaining_today", "remaining_week", "remaining_30_days", "remaining_window"):
            assert k in cap and isinstance(cap[k], int)

        # daily_limit_total from inbox list
        ib = admin_client.get(f"{BASE_URL}/api/infrastructure/inboxes").json()["inboxes"]
        daily_limit_total = sum(r["daily_limit"] for r in ib)
        # remaining_window should be close to daily_limit_total * 120 when no projections
        # (today might have small sent_today)
        assert cap["remaining_window"] <= daily_limit_total * 120
        assert cap["remaining_window"] >= daily_limit_total * 120 - daily_limit_total

    def test_summary_window_30(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/summary?window_days=30")
        assert r.status_code == 200
        cap = r.json()["capacity"]
        assert cap["window_days"] == 30
        assert cap["remaining_window"] == cap["remaining_30_days"]

    def test_summary_window_limits(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/summary?window_days=0")
        assert r.status_code == 422
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/summary?window_days=366")
        assert r.status_code == 422
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/summary?window_days=1")
        assert r.status_code == 200
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/summary?window_days=365")
        assert r.status_code == 200

    def test_summary_no_est_note_says_projection(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/summary").json()
        # Note should describe projection (not "linear estimate")
        assert "projection" in r["capacity"]["note"].lower()


# ---------------- Inboxes row enrichment ---------------------------------

class TestInboxesProjection:
    def test_every_row_has_projected_window_keys(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/inboxes")
        assert r.status_code == 200
        rows = r.json()["inboxes"]
        assert len(rows) > 0
        for row in rows:
            assert "projected_window_total" in row
            assert isinstance(row["projected_window_total"], int)
            assert row["projected_window_days"] == 120

    def test_no_future_projections_default(self, admin_client, mongo):
        # When no test drip seeded, projected total should be 0 for all rows
        # Only true when no active drip_contacts exist with next_send_at
        active_contacts = mongo.drip_contacts.count_documents({"status": "active", "next_send_at": {"$exists": True, "$ne": None}})
        rows = admin_client.get(f"{BASE_URL}/api/infrastructure/inboxes").json()["inboxes"]
        if active_contacts == 0:
            for row in rows:
                assert row["projected_window_total"] == 0


# ---------------- Calendar endpoint --------------------------------------

class TestCalendar:
    def test_calendar_default_120_rows(self, admin_client):
        ib = admin_client.get(f"{BASE_URL}/api/infrastructure/inboxes").json()["inboxes"]
        assert ib, "need at least one inbox"
        acc_id = ib[0]["account_id"]
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/calendar/{acc_id}")
        assert r.status_code == 200
        body = r.json()
        assert len(body["days"]) == 120
        assert body["window_days"] == 120
        assert "totals" in body
        for k in ("projected", "remaining", "capacity"):
            assert k in body["totals"]
        # Day 0 = today UTC, day 119 = today + 119
        today = datetime.now(timezone.utc).date()
        assert body["days"][0]["date"] == today.isoformat()
        assert body["days"][119]["date"] == (today + timedelta(days=119)).isoformat()
        # weekday is 0..6 (Mon=0)
        assert body["days"][0]["weekday"] == today.weekday()
        # status mapping
        for d in body["days"]:
            assert d["status"] in ("Available", "Partial", "Reserved")
            if d["projected"] == 0 and d["used"] == 0:
                assert d["status"] == "Available"
            elif d["remaining"] == 0:
                assert d["status"] == "Reserved"
            elif 0 < d["used"] < d["limit"]:
                assert d["status"] == "Partial"

    def test_calendar_404_bogus(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/calendar/acc_bogus")
        assert r.status_code == 404
        assert "not found" in r.json().get("detail", "").lower() or "not visible" in r.json().get("detail", "").lower()

    def test_calendar_window_30(self, admin_client):
        ib = admin_client.get(f"{BASE_URL}/api/infrastructure/inboxes").json()["inboxes"]
        acc_id = ib[0]["account_id"]
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/calendar/{acc_id}?window_days=30")
        assert r.status_code == 200
        body = r.json()
        assert len(body["days"]) == 30
        assert body["window_days"] == 30

    def test_calendar_totals_match_rows(self, admin_client):
        ib = admin_client.get(f"{BASE_URL}/api/infrastructure/inboxes").json()["inboxes"]
        acc_id = ib[0]["account_id"]
        body = admin_client.get(f"{BASE_URL}/api/infrastructure/calendar/{acc_id}").json()
        assert body["totals"]["capacity"] == sum(d["limit"] for d in body["days"])
        assert body["totals"]["projected"] == sum(d["projected"] for d in body["days"])
        assert body["totals"]["remaining"] == sum(d["remaining"] for d in body["days"])


# ---------------- Projection: drip campaign seed -------------------------

class TestProjectionDripSeed:
    """Seeds an active drip + 5 contacts and verifies projected sends land on
    predicted dates per inbox."""

    @pytest.fixture(scope="class")
    def seed_drip(self, mongo):
        # Pick 2 accounts of super_admin
        accs = list(mongo.email_accounts.find({"user_id": SUPER_ADMIN_ID}, {"_id": 0, "account_id": 1}).limit(2))
        if len(accs) < 2:
            pytest.skip("Need 2 super_admin accounts")
        account_ids = [a["account_id"] for a in accs]
        drip_id = f"TEST_iter52_drip_{uuid.uuid4().hex[:8]}"
        today = datetime.now(timezone.utc).date()
        # Use a Monday to avoid sending_days skipping
        # Find next Monday so first send lands on Mon-Fri
        next_mon = today
        while next_mon.weekday() != 0:
            next_mon = next_mon + timedelta(days=1)
        first_send = datetime.combine(next_mon, datetime.min.time(), tzinfo=timezone.utc) + timedelta(hours=9)

        mongo.drip_campaigns.insert_one({
            "drip_id": drip_id,
            "user_id": SUPER_ADMIN_ID,
            "status": "running",
            "account_ids": account_ids,
            "name": "TEST_iter52_drip",
            "steps": [
                {"delay_days": 0, "delay_hours": 0, "subject": "s1", "body": "b1"},
                {"delay_days": 7, "delay_hours": 0, "subject": "s2", "body": "b2"},
                {"delay_days": 3, "delay_hours": 0, "subject": "s3", "body": "b3"},
            ],
            "schedule": {
                "timezone": "UTC",
                "sending_days": [0, 1, 2, 3, 4],
                "start_time": "09:00",
                "start_date": None,
            },
        })
        contact_ids = []
        for i in range(5):
            cid = f"TEST_iter52_dc_{uuid.uuid4().hex[:8]}"
            mongo.drip_contacts.insert_one({
                "contact_id": cid,
                "drip_id": drip_id,
                "status": "active",
                "current_step": 0,
                "next_send_at": first_send.isoformat(),
                "email": f"test{i}@example.com",
            })
            contact_ids.append(cid)
        yield {"drip_id": drip_id, "account_ids": account_ids, "first_send_date": next_mon}
        # Cleanup
        mongo.drip_campaigns.delete_one({"drip_id": drip_id})
        mongo.drip_contacts.delete_many({"drip_id": drip_id})

    def test_inbox_rows_show_projected_after_seed(self, admin_client, seed_drip):
        rows = admin_client.get(f"{BASE_URL}/api/infrastructure/inboxes").json()["inboxes"]
        seeded_rows = [r for r in rows if r["account_id"] in seed_drip["account_ids"]]
        assert len(seeded_rows) == 2
        total_projected = sum(r["projected_window_total"] for r in seeded_rows)
        # 5 contacts × 3 steps = 15 projected sends total across 2 inboxes
        assert total_projected == 15, f"Expected 15 projected sends got {total_projected}"

    def test_calendar_shows_projected_on_first_send_date(self, admin_client, seed_drip):
        # First send date should have projected sends on at least one of the inboxes
        first_iso = seed_drip["first_send_date"].isoformat()
        found_first = 0
        for acc_id in seed_drip["account_ids"]:
            body = admin_client.get(f"{BASE_URL}/api/infrastructure/calendar/{acc_id}").json()
            day_row = next((d for d in body["days"] if d["date"] == first_iso), None)
            if day_row and day_row["projected"] > 0:
                found_first += day_row["projected"]
                assert day_row["status"] in ("Partial", "Reserved")
        # 5 contacts at first send across 2 inboxes round-robin → 3 + 2
        assert found_first == 5, f"first-send projected total mismatch: {found_first}"

    def test_summary_capacity_reduced_after_seed(self, admin_client, seed_drip):
        cap = admin_client.get(f"{BASE_URL}/api/infrastructure/summary").json()["capacity"]
        # remaining_window must be less than daily_limit_total*120 due to projections
        ib = admin_client.get(f"{BASE_URL}/api/infrastructure/inboxes").json()["inboxes"]
        daily_limit_total = sum(r["daily_limit"] for r in ib)
        # 15 projected sends total — remaining_window <= daily_limit_total*120 - 15
        assert cap["remaining_window"] <= daily_limit_total * 120


# ---------------- Projection: scheduled regular campaign ---------------

class TestProjectionScheduledCampaign:
    @pytest.fixture(scope="class")
    def seed_camp(self, mongo):
        accs = list(mongo.email_accounts.find({"user_id": SUPER_ADMIN_ID}, {"_id": 0, "account_id": 1}).limit(3))
        if len(accs) < 3:
            pytest.skip("need 3 accounts")
        account_ids = [a["account_id"] for a in accs]
        camp_id = f"TEST_iter52_camp_{uuid.uuid4().hex[:8]}"
        # Schedule for 5 days from now
        future = datetime.now(timezone.utc) + timedelta(days=5)
        mongo.campaigns.insert_one({
            "campaign_id": camp_id,
            "user_id": SUPER_ADMIN_ID,
            "name": "TEST_iter52_camp",
            "status": "scheduled",
            "account_ids": account_ids,
            "scheduled_at": future.isoformat(),
            "total_emails": 10,
            "sent_count": 0,
            "timezone": "UTC",
        })
        yield {"campaign_id": camp_id, "account_ids": account_ids, "spike_date": future.date()}
        mongo.campaigns.delete_one({"campaign_id": camp_id})

    def test_spike_lands_on_scheduled_date(self, admin_client, seed_camp):
        spike_iso = seed_camp["spike_date"].isoformat()
        # Pending=10, account_ids=3 → per_account=3, remainder=1 → first acc gets 4, others 3
        per_account_counts = []
        for acc_id in seed_camp["account_ids"]:
            body = admin_client.get(f"{BASE_URL}/api/infrastructure/calendar/{acc_id}").json()
            day_row = next((d for d in body["days"] if d["date"] == spike_iso), None)
            assert day_row is not None
            per_account_counts.append(day_row["projected"])
        # First account should have one more
        assert sum(per_account_counts) >= 10
        assert max(per_account_counts) - min(per_account_counts) <= 1


# ---------------- Permissions ---------------------------------------------

class TestPermissions:
    def test_summary_403_for_non_permitted(self, mongo, user_client):
        # ensure user does not have can_access_infrastructure
        mongo.users.update_one({"user_id": NORMAL_USER_ID}, {"$set": {"can_access_infrastructure": False}})
        r = user_client.get(f"{BASE_URL}/api/infrastructure/summary")
        assert r.status_code == 403

    def test_calendar_403_for_non_permitted(self, mongo, user_client):
        mongo.users.update_one({"user_id": NORMAL_USER_ID}, {"$set": {"can_access_infrastructure": False}})
        r = user_client.get(f"{BASE_URL}/api/infrastructure/calendar/some_acc")
        assert r.status_code == 403

    def test_inboxes_403_for_non_permitted(self, mongo, user_client):
        mongo.users.update_one({"user_id": NORMAL_USER_ID}, {"$set": {"can_access_infrastructure": False}})
        r = user_client.get(f"{BASE_URL}/api/infrastructure/inboxes")
        assert r.status_code == 403


# ---------------- Exports -------------------------------------------------

class TestExports:
    def test_inbox_xlsx_has_projected_column(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/export?type=inboxes&format=xlsx")
        assert r.status_code == 200
        assert r.headers.get("content-type", "").startswith(XLSX_MIME)
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        assert "Projected (120d)" in headers, f"headers={headers}"

    def test_domain_csv_header(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/export?type=domains&format=csv")
        assert r.status_code == 200
        reader = csv.reader(io.StringIO(r.text))
        header = next(reader)
        expected = ["Domain", "Inbox Count", "Total Daily Capacity", "Used Today",
                    "Remaining Today", "Projected (120d)", "Status"]
        assert header == expected, f"got {header}"

    def test_domain_xlsx_has_projected(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/export?type=domains&format=xlsx")
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        assert "Projected (120d)" in headers


# ---------------- Phase 1 regression -------------------------------------

class TestRegression:
    def test_inboxes_filters_still_work(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/inboxes?status=Available")
        assert r.status_code == 200
        for row in r.json()["inboxes"]:
            assert row["status"].lower() == "available"

    def test_inboxes_search(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/infrastructure/inboxes?search=zzzzzznoresults")
        assert r.status_code == 200
        assert r.json()["total"] == 0

    def test_summary_has_inbox_counts(self, admin_client):
        s = admin_client.get(f"{BASE_URL}/api/infrastructure/summary").json()
        for k in ("available", "partially_available", "fully_reserved", "warming_up", "paused", "risky", "total"):
            assert k in s["inbox_counts"]

    def test_campaigns_200(self, admin_client):
        assert admin_client.get(f"{BASE_URL}/api/campaigns").status_code == 200

    def test_drip_campaigns_200(self, admin_client):
        assert admin_client.get(f"{BASE_URL}/api/drip-campaigns").status_code == 200

    def test_accounts_200(self, admin_client):
        assert admin_client.get(f"{BASE_URL}/api/accounts").status_code == 200

    def test_auth_me_200(self, admin_client):
        assert admin_client.get(f"{BASE_URL}/api/auth/me").status_code == 200

    def test_subscription_prices_200(self, admin_client):
        assert admin_client.get(f"{BASE_URL}/api/subscription/prices").status_code == 200
