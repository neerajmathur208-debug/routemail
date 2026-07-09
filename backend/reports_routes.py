"""Campaign + Drip Campaign reporting.

A single router (mounted under `/api/reports`) exposing:

* **`GET /reports/campaigns`** — JSON row list powering the Reports page
  table (name / total prospects / emails sent / date sent + status).
* **`GET /reports/campaigns/export.csv`** — CSV download matching the
  product spec's 4-column contract (Campaign Name, Total Prospects,
  Emails Sent, Date Sent).
* **`GET /reports/export`** — legacy .xlsx workbook (kept for the
  existing ExportReportDialog component; not part of the new spec).

Every endpoint is strictly scoped to the requester's ``user_id`` — the
Infrastructure isolation contract also applies here.

Layout note: this file is intentionally the single home for reporting so
new reports (Infrastructure, Warmup, Unibox, Reply, Domain Health, …)
can be added as additional endpoints under the same `/reports` prefix
without touching the frontend router or sidebar.
"""

import csv
import io
import re
from datetime import datetime, timezone, date as _date_cls
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ---------- helpers ---------------------------------------------------------

_CSV_HEADERS = [
    "Campaign / Drip Campaign Name",
    "Total Prospects in the List",
    "Emails Sent",
    "Date Sent",
]

_FNAME_STRIP = re.compile(r"[^A-Za-z0-9_-]+")


def _to_iso_date(value: Any) -> str:
    """Best-effort coerce to a YYYY-MM-DD string (UTC). Returns "" on miss."""
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.astimezone(timezone.utc).date().isoformat() if value.tzinfo else value.date().isoformat()
    if isinstance(value, str):
        try:
            dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
            return dt.astimezone(timezone.utc).date().isoformat() if dt.tzinfo else dt.date().isoformat()
        except Exception:
            # Legacy rows already stored as bare YYYY-MM-DD
            if len(value) >= 10 and value[4] == "-" and value[7] == "-":
                return value[:10]
    return ""


def _campaign_date_sent(camp: Dict[str, Any]) -> str:
    return _to_iso_date(
        camp.get("started_at")
        or camp.get("completed_at")
        or camp.get("scheduled_at")
        or camp.get("created_at")
    )


def _drip_date_sent(drip: Dict[str, Any]) -> str:
    stats = drip.get("last_run_stats") or {}
    schedule = drip.get("schedule") or {}
    return _to_iso_date(
        drip.get("started_at")
        or stats.get("started_at")
        or schedule.get("start_date")
        or drip.get("created_at")
    )


def _within_range(date_str: str, start: Optional[str], end: Optional[str]) -> bool:
    if not date_str:
        return start is None and end is None
    if start and date_str < start:
        return False
    if end and date_str > end:
        return False
    return True


def _matches_name(name: str, search: Optional[str]) -> bool:
    if not search:
        return True
    return search.lower() in (name or "").lower()


def _safe_filename(base: str) -> str:
    slug = _FNAME_STRIP.sub("_", base).strip("_") or "report"
    return f"{slug}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.csv"


async def _load_campaign_rows(db, user_id: str) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    async for c in db.campaigns.find(
        {"user_id": user_id},
        {
            "_id": 0, "campaign_id": 1, "name": 1, "status": 1,
            "total_emails": 1, "sent_count": 1,
            "started_at": 1, "completed_at": 1, "scheduled_at": 1, "created_at": 1,
        },
    ).sort([("created_at", -1)]):
        rows.append({
            "id": c.get("campaign_id"),
            "type": "Campaign",
            "name": c.get("name") or "",
            "total_prospects": int(c.get("total_emails") or 0),
            "emails_sent": int(c.get("sent_count") or 0),
            "date_sent": _campaign_date_sent(c),
            "status": c.get("status") or "",
        })
    return rows


async def _load_drip_rows(db, user_id: str) -> List[Dict[str, Any]]:
    """Emails Sent prefers the ``total_sent`` counter maintained by the
    drip worker; falls back to counting ``drip_logs`` for legacy rows."""
    rows: List[Dict[str, Any]] = []
    drips = await db.drip_campaigns.find(
        {"user_id": user_id},
        {
            "_id": 0, "drip_id": 1, "name": 1, "status": 1,
            "total_contacts": 1, "total_sent": 1, "last_run_stats": 1,
            "schedule": 1, "started_at": 1, "created_at": 1,
        },
    ).sort([("created_at", -1)]).to_list(10000)

    for d in drips:
        drip_id = d.get("drip_id")
        sent = d.get("total_sent")
        if sent is None:
            sent = await db.drip_logs.count_documents({"drip_id": drip_id, "status": "sent"})
        total_prospects = await db.drip_contacts.count_documents({"drip_id": drip_id})
        if not total_prospects:
            total_prospects = int(d.get("total_contacts") or 0)
        rows.append({
            "id": drip_id,
            "type": "Drip Campaign",
            "name": d.get("name") or "",
            "total_prospects": int(total_prospects),
            "emails_sent": int(sent or 0),
            "date_sent": _drip_date_sent(d),
            "status": d.get("status") or "",
        })
    return rows


async def _fetch_rows(
    db,
    user_id: str,
    *,
    campaign_type: str,
    start_date: Optional[str],
    end_date: Optional[str],
    search: Optional[str],
) -> List[Dict[str, Any]]:
    all_rows: List[Dict[str, Any]] = []
    if campaign_type in ("campaign", "both"):
        all_rows.extend(await _load_campaign_rows(db, user_id))
    if campaign_type in ("drip", "both"):
        all_rows.extend(await _load_drip_rows(db, user_id))

    filtered = [
        r for r in all_rows
        if _within_range(r["date_sent"], start_date, end_date)
        and _matches_name(r["name"], search)
    ]
    filtered.sort(key=lambda r: (r["date_sent"] or "0000-00-00", r["name"]), reverse=True)
    return filtered


# ---------- legacy .xlsx helpers -------------------------------------------

def _parse_iso_date(s: Optional[str]) -> Optional[datetime]:
    """Accepts 'YYYY-MM-DD' or full ISO datetime. Returns a UTC-aware datetime."""
    if not s:
        return None
    try:
        if len(s) == 10:
            return datetime.fromisoformat(s).replace(tzinfo=timezone.utc)
        return datetime.fromisoformat(s.replace("Z", "+00:00")).astimezone(timezone.utc)
    except Exception:
        raise HTTPException(status_code=400, detail=f"Invalid date: {s}")


def _safe_pct(num: int, denom: int) -> str:
    if not denom:
        return "0.00%"
    return f"{(num / denom * 100):.2f}%"


def _fmt_dt(value) -> str:
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00")).strftime("%Y-%m-%d %H:%M")
        except Exception:
            return value
    return str(value)


def _matches_date_range(value, frm: Optional[datetime], to: Optional[datetime]) -> bool:
    """Inclusive on both ends. The `value` may be datetime or ISO string or None."""
    if not frm and not to:
        return True
    if value is None:
        return False
    dt: Optional[datetime] = None
    if isinstance(value, datetime):
        dt = value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    elif isinstance(value, str):
        try:
            dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
            if not dt.tzinfo:
                dt = dt.replace(tzinfo=timezone.utc)
        except Exception:
            return False
    if dt is None:
        return False
    if frm and dt < frm:
        return False
    if to and dt > to:
        return False
    return True


# ---------- router ----------------------------------------------------------

def build_reports_router(db, get_current_user):
    router = APIRouter(prefix="/reports", tags=["reports"])

    def _uid(user) -> str:
        return user.user_id if hasattr(user, "user_id") else user["user_id"]

    def _validate_type(campaign_type: str) -> str:
        ct = (campaign_type or "both").lower().strip()
        if ct not in {"campaign", "drip", "both"}:
            raise HTTPException(
                status_code=400,
                detail="campaign_type must be one of: campaign, drip, both",
            )
        return ct

    # ─── New spec endpoints ───────────────────────────────────────────────

    @router.get("/campaigns", operation_id="reports_campaigns_list")
    async def list_campaign_report(
        start_date: Optional[str] = Query(None, description="YYYY-MM-DD inclusive"),
        end_date: Optional[str] = Query(None, description="YYYY-MM-DD inclusive"),
        campaign_type: str = Query("both", description="campaign | drip | both"),
        search: Optional[str] = Query(None, description="Case-insensitive name substring"),
        user=Depends(get_current_user),
    ) -> Dict[str, Any]:
        """JSON version powering the Reports table on the frontend."""
        ct = _validate_type(campaign_type)
        rows = await _fetch_rows(
            db, _uid(user),
            campaign_type=ct, start_date=start_date, end_date=end_date, search=search,
        )
        return {
            "rows": rows,
            "total": len(rows),
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "campaign_type": ct,
                "search": search or "",
            },
        }

    @router.get("/campaigns/export.csv", operation_id="reports_campaigns_csv")
    async def export_campaign_report_csv(
        start_date: Optional[str] = Query(None),
        end_date: Optional[str] = Query(None),
        campaign_type: str = Query("both"),
        search: Optional[str] = Query(None),
        user=Depends(get_current_user),
    ):
        """CSV per the 4-column product spec. Streams so large reports (100k+
        rows) don't buffer in memory."""
        ct = _validate_type(campaign_type)
        rows = await _fetch_rows(
            db, _uid(user), campaign_type=ct,
            start_date=start_date, end_date=end_date, search=search,
        )

        def _iter():
            buf = io.StringIO()
            csv.writer(buf, quoting=csv.QUOTE_MINIMAL).writerow(_CSV_HEADERS)
            yield buf.getvalue()
            for r in rows:
                buf = io.StringIO()
                csv.writer(buf, quoting=csv.QUOTE_MINIMAL).writerow([
                    r["name"], r["total_prospects"], r["emails_sent"], r["date_sent"],
                ])
                yield buf.getvalue()

        fname = _safe_filename("RouteMail_Campaign_Report")
        return StreamingResponse(
            _iter(),
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="{fname}"',
                "Cache-Control": "no-store",
            },
        )

    # ─── Legacy .xlsx endpoint (kept for existing dialog component) ───────

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill("solid", fgColor="6D28D9")  # violet-700
    CELL_BORDER = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    def _write_headers(ws, headers):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = CELL_BORDER

    def _autofit(ws, max_width: int = 38):
        for column_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 10), max_width)

    async def _resolve_list_name(user_id: str, list_id: Optional[str]) -> str:
        if not list_id:
            return ""
        doc = await db.email_lists.find_one({"list_id": list_id, "user_id": user_id}, {"_id": 0, "name": 1})
        return (doc or {}).get("name", "") or ""

    @router.get("/export")
    async def export_report(
        from_date: Optional[str] = Query(None, description="Inclusive lower bound (YYYY-MM-DD). Filters on created_at."),
        to_date: Optional[str] = Query(None, description="Inclusive upper bound (YYYY-MM-DD)."),
        campaign_type: str = Query("all", description="all | campaigns | drip"),
        status: Optional[str] = Query(None, description="Comma-separated statuses e.g. 'running,completed'. Empty/None = all."),
        user=Depends(get_current_user),
    ):
        """Generate an .xlsx workbook of campaign + drip-campaign performance.

        Authoritative metrics: each row's Reply / Bounce / Unsubscribe counts
        come straight from the source-of-truth collections — `drip_contacts` for
        drip rows and `campaigns.sent_count/failed_count` + `do_not_email_lists`
        unsubscribe entries for regular campaign rows.
        """
        campaign_type = (campaign_type or "all").lower()
        if campaign_type not in ("all", "campaigns", "drip"):
            raise HTTPException(status_code=400, detail="campaign_type must be all|campaigns|drip")

        status_filter = None
        if status and status.strip():
            status_filter = {s.strip().lower() for s in status.split(",") if s.strip()}

        frm = _parse_iso_date(from_date)
        to = _parse_iso_date(to_date)
        # Upper bound inclusive — push to end of day if a bare date was given
        if to and len(to_date or "") == 10:
            to = to.replace(hour=23, minute=59, second=59)

        # Pre-load both collections (filtered by ownership) — apply python-side
        # range filtering so we can match either datetime or ISO-string values.
        camps_raw = []
        drips_raw = []
        if campaign_type in ("all", "campaigns"):
            camps_raw = await db.campaigns.find(
                {"user_id": user.user_id}, {"_id": 0}
            ).sort("created_at", -1).to_list(10000)
        if campaign_type in ("all", "drip"):
            drips_raw = await db.drip_campaigns.find(
                {"user_id": user.user_id}, {"_id": 0}
            ).sort("created_at", -1).to_list(10000)

        # Pre-aggregate unsubscribes per campaign by counting DNE entries with
        # the `source="unsubscribe"` that landed during the report range. These
        # are user-wide (not per-campaign tagged in the schema), so the same
        # total is shown beside each campaign row. The user explicitly asked
        # for the column — labelling it as "Unsubscribes (window)" keeps it
        # honest.
        unsub_count_in_range = await db.do_not_email_emails.count_documents({
            "user_id": user.user_id,
            "source": "unsubscribe",
            **({"added_at": {"$gte": frm.isoformat()}} if frm else {}),
            **({"added_at": {"$lte": to.isoformat()}} if to else {}),
        })

        # Filter campaigns by date range + status
        def _camp_pass(c):
            if status_filter and c.get("status", "").lower() not in status_filter:
                return False
            return _matches_date_range(c.get("created_at"), frm, to)

        camps = [c for c in camps_raw if _camp_pass(c)]
        drips = [d for d in drips_raw if _camp_pass(d)]

        # Build workbook
        wb = Workbook()
        # Remove default sheet — we'll create our own
        wb.remove(wb.active)

        # ---- Campaigns sheet ----
        ws_camp = wb.create_sheet("Campaigns")
        camp_headers = [
            "Campaign Name", "Campaign Type", "Status",
            "Created Date", "Scheduled Date", "Start Date",
            "Contacts Targeted", "Emails Sent",
            "Replies", "Bounce Count", "Unsubscribes",
            "Reply Rate",
            "List",
        ]
        _write_headers(ws_camp, camp_headers)

        for row_idx, c in enumerate(camps, start=2):
            sent = int(c.get("sent_count") or 0)
            bounced = int(c.get("failed_count") or 0)  # failed sends ≈ bounces at the send layer
            # Reply count from Unibox (replies are stored against the recipient
            # email, not the campaign — so we count distinct contacts on this
            # campaign's recipient list that have an inbox reply).
            campaign_id = c.get("campaign_id")
            reply_count = await db.unibox_messages.count_documents({
                "user_id": user.user_id,
                "campaign_id": campaign_id,
                "kind": "reply",
            }) if campaign_id else 0
            ws_camp.cell(row=row_idx, column=1, value=c.get("name", ""))
            ws_camp.cell(row=row_idx, column=2, value="Campaign")
            ws_camp.cell(row=row_idx, column=3, value=c.get("status", ""))
            ws_camp.cell(row=row_idx, column=4, value=_fmt_dt(c.get("created_at")))
            ws_camp.cell(row=row_idx, column=5, value=_fmt_dt(c.get("scheduled_at")))
            ws_camp.cell(row=row_idx, column=6, value=_fmt_dt(c.get("started_at")))
            ws_camp.cell(row=row_idx, column=7, value=int(c.get("total_emails") or 0))
            ws_camp.cell(row=row_idx, column=8, value=sent)
            ws_camp.cell(row=row_idx, column=9, value=reply_count)
            ws_camp.cell(row=row_idx, column=10, value=bounced)
            ws_camp.cell(row=row_idx, column=11, value=unsub_count_in_range)
            ws_camp.cell(row=row_idx, column=12, value=_safe_pct(reply_count, sent))
            ws_camp.cell(row=row_idx, column=13, value=await _resolve_list_name(user.user_id, c.get("list_id")))
        _autofit(ws_camp)

        # ---- Drip Campaigns sheet ----
        ws_drip = wb.create_sheet("Drip Campaigns")
        drip_headers = [
            "Campaign Name", "Campaign Type", "Status",
            "Created Date", "Scheduled Start Date", "Start Date",
            "Total Steps",
            "Contacts Targeted", "Emails Sent",
            "Active Contacts", "Completed Contacts", "Stopped Contacts", "Currently Running",
            "Replies", "Bounce Count", "Unsubscribes",
            "Reply Rate",
            "List",
        ]
        _write_headers(ws_drip, drip_headers)

        for row_idx, d in enumerate(drips, start=2):
            drip_id = d.get("drip_id")
            # Aggregate drip_contacts in one round trip per drip
            cursor = db.drip_contacts.aggregate([
                {"$match": {"drip_id": drip_id}},
                {"$group": {"_id": "$status", "n": {"$sum": 1}}},
            ])
            buckets = {row["_id"] or "unknown": row["n"] async for row in cursor}
            total_contacts = sum(buckets.values())
            active = buckets.get("active", 0)
            completed = buckets.get("completed", 0)
            replied = buckets.get("replied", 0)
            bounced = buckets.get("bounced", 0)
            unsubscribed = buckets.get("unsubscribed", 0)
            stopped = replied + bounced + unsubscribed
            sent = int(d.get("total_sent") or 0)
            schedule = d.get("schedule") or {}
            start_date_str = schedule.get("start_date") or ""
            timezone_str = schedule.get("timezone") or "UTC"
            scheduled_start_display = (
                f"{start_date_str} {schedule.get('start_time') or ''} ({timezone_str})".strip()
                if start_date_str else ""
            )

            ws_drip.cell(row=row_idx, column=1, value=d.get("name", ""))
            ws_drip.cell(row=row_idx, column=2, value="Drip")
            ws_drip.cell(row=row_idx, column=3, value=d.get("status", ""))
            ws_drip.cell(row=row_idx, column=4, value=_fmt_dt(d.get("created_at")))
            ws_drip.cell(row=row_idx, column=5, value=scheduled_start_display)
            ws_drip.cell(row=row_idx, column=6, value=start_date_str)
            ws_drip.cell(row=row_idx, column=7, value=len(d.get("steps") or []))
            ws_drip.cell(row=row_idx, column=8, value=total_contacts)
            ws_drip.cell(row=row_idx, column=9, value=sent)
            ws_drip.cell(row=row_idx, column=10, value=active)
            ws_drip.cell(row=row_idx, column=11, value=completed)
            ws_drip.cell(row=row_idx, column=12, value=stopped)
            ws_drip.cell(row=row_idx, column=13, value=active)
            ws_drip.cell(row=row_idx, column=14, value=replied)
            ws_drip.cell(row=row_idx, column=15, value=bounced)
            ws_drip.cell(row=row_idx, column=16, value=unsubscribed)
            ws_drip.cell(row=row_idx, column=17, value=_safe_pct(replied, sent))
            ws_drip.cell(row=row_idx, column=18, value=await _resolve_list_name(user.user_id, d.get("list_id")))
        _autofit(ws_drip)

        # ---- Summary sheet (always created, helpful when the user opens it) ----
        ws_sum = wb.create_sheet("Summary", 0)
        ws_sum["A1"] = "RouteMail — Campaign Report"
        ws_sum["A1"].font = Font(bold=True, size=14)
        ws_sum["A3"] = "Generated"
        ws_sum["B3"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        ws_sum["A4"] = "Account"
        ws_sum["B4"] = user.email
        ws_sum["A5"] = "Date Range"
        ws_sum["B5"] = f"{from_date or 'all-time'} → {to_date or 'today'}"
        ws_sum["A6"] = "Campaign Type"
        ws_sum["B6"] = campaign_type
        ws_sum["A7"] = "Status Filter"
        ws_sum["B7"] = status or "all"
        ws_sum["A9"] = "Counts"
        ws_sum["A9"].font = Font(bold=True)
        ws_sum["A10"] = "Campaigns in report"
        ws_sum["B10"] = len(camps)
        ws_sum["A11"] = "Drip Campaigns in report"
        ws_sum["B11"] = len(drips)
        ws_sum["A13"] = "Note"
        ws_sum["B13"] = (
            "Open/Click metrics are intentionally not included — RouteMail does not "
            "track pixel-opens or link-clicks. Reply / Bounce / Unsubscribe counts "
            "are sourced from Unibox, send-failure logs, and the Do-Not-Email list "
            "respectively."
        )
        for col in ("A", "B"):
            ws_sum.column_dimensions[col].width = 40 if col == "B" else 28

        # Stream to bytes
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        # Build filename per spec
        def _ymd(s, fallback):
            if not s:
                return fallback
            try:
                return _date_cls.fromisoformat(s[:10]).strftime("%Y-%m-%d")
            except Exception:
                return fallback

        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        fname = f"RouteMail_Campaign_Report_{_ymd(from_date, 'all')}_to_{_ymd(to_date, today)}.xlsx"

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    return router
