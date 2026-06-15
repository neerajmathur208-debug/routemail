"""Infrastructure Phase A — Email-account export + Forecasting + Domain tracking + Renewal alerts.

Single router mounted at /api/infrastructure. All four endpoints are gated by
the existing `get_infrastructure_user` dep (super_admin OR can_access_infrastructure).
"""
import csv as _csv
import io
import uuid
from datetime import datetime, timedelta, timezone, date as _date_cls
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, Field

from infra_projection import build_projection, aggregate_capacity


async def ensure_domain_record(db, user_id: str, email: str) -> Optional[str]:
    """Phase Batch-1 auto-detection — when an email account is added, ensure
    a `tracked_domains` row exists for the recipient domain.

    Rules (confirmed with user):
      * Silent — no notification.
      * Idempotent — only one row per (user_id, domain). Existing rows are
        NEVER overwritten; only the linked_inbox_count is incremented.
      * Default purchase_date / date_added = today (date the inbox arrived
        in RouteMail).
      * Default expiry_date = today + 361 days (per spec).
      * Default renewal_date = expiry_date.
    """
    if not email or "@" not in email:
        return None
    domain = email.rsplit("@", 1)[1].strip().lower()
    if not domain or "." not in domain:
        return None
    existing = await db.tracked_domains.find_one(
        {"user_id": user_id, "domain": domain}, {"_id": 0}
    )
    today = _date_cls.today()
    if existing:
        await db.tracked_domains.update_one(
            {"user_id": user_id, "domain": domain},
            {"$inc": {"linked_inbox_count": 1}},
        )
        return existing.get("domain_id")
    domain_id = f"dom_{uuid.uuid4().hex[:10]}"
    await db.tracked_domains.insert_one({
        "domain_id": domain_id,
        "user_id": user_id,
        "domain": domain,
        "registrar": None,
        "purchase_date": today.isoformat(),
        "date_added": today.isoformat(),
        "expiry_date": (today + timedelta(days=361)).isoformat(),
        "renewal_date": (today + timedelta(days=361)).isoformat(),
        "notes": None,
        "linked_inbox_count": 1,
        "auto_created": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    return domain_id



def _xlsx(headers: List[str], rows: List[List[Any]], sheet: str, fill: str = "4338CA") -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = sheet
    font = Font(bold=True, color="FFFFFF"); pf = PatternFill("solid", fgColor=fill)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h); c.font = font; c.fill = pf
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    for col in ws.columns:
        ln = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(ln + 2, 10), 40)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()


def _csv_bytes(headers: List[str], rows: List[List[Any]]) -> bytes:
    out = io.StringIO(); w = _csv.writer(out); w.writerow(headers)
    for r in rows: w.writerow(r)
    return out.getvalue().encode()


class DomainUpsert(BaseModel):
    domain: str = Field(..., min_length=3, max_length=253)
    registrar: Optional[str] = Field(None, max_length=120)
    purchase_date: Optional[str] = None
    expiry_date: Optional[str] = None
    renewal_date: Optional[str] = None
    date_added: Optional[str] = None
    notes: Optional[str] = Field(None, max_length=500)


def attach_phase_a_routes(router: APIRouter, db, get_infra_user, load_inboxes_fn):

    # ───────────── 1. EXPORT EMAIL ACCOUNTS ─────────────
    @router.get("/accounts/export")
    async def export_accounts(format: str = Query("xlsx"), user=Depends(get_infra_user)):
        fmt = (format or "xlsx").lower()
        if fmt not in ("xlsx", "csv"):
            raise HTTPException(status_code=400, detail="format must be xlsx|csv")
        is_admin = user.get("role") == "super_admin"
        q = {} if is_admin else {"user_id": user["user_id"]}
        accts = await db.email_accounts.find(q, {"_id": 0}).sort("email", 1).to_list(10000)

        # Index active campaign + drip assignments per account_id
        from collections import defaultdict
        assign = defaultdict(list)
        camp_q = {"status": {"$in": ["running", "scheduled", "paused", "paused_daily_limit"]}}
        if not is_admin: camp_q["user_id"] = user["user_id"]
        for c in await db.campaigns.find(camp_q, {"_id": 0, "name": 1, "account_ids": 1}).to_list(5000):
            for aid in c.get("account_ids") or []: assign[aid].append(c.get("name") or "")
        drip_q = {"status": {"$in": ["running", "scheduled", "paused"]}}
        if not is_admin: drip_q["user_id"] = user["user_id"]
        for d in await db.drip_campaigns.find(drip_q, {"_id": 0, "name": 1, "account_ids": 1}).to_list(5000):
            for aid in d.get("account_ids") or []: assign[aid].append(f"[drip] {d.get('name') or ''}")

        headers = ["Email", "Domain", "Ownership", "SMTP Host", "SMTP Port", "SMTP Username",
                   "IMAP Host", "IMAP Port", "IMAP Username", "Daily Limit", "Status",
                   "Warmup Status", "Last Activity", "Date Added", "Campaign Assignments", "Notes"]
        rows = []
        for a in accts:
            email = a.get("email", "")
            rows.append([
                email, email.rsplit("@", 1)[-1] if "@" in email else "",
                a.get("ownership") or "",
                a.get("smtp_host") or "", a.get("smtp_port") or "", a.get("smtp_username") or "",
                a.get("imap_host") or "", a.get("imap_port") or "", a.get("imap_username") or "",
                int(a.get("daily_limit") or 50),
                (a.get("status") or "").lower(),
                (a.get("warmup_status") or "—") if a.get("warmup_enabled") else "—",
                a.get("last_sent_at") or "",
                a.get("created_at") or a.get("added_at") or "",
                ", ".join(assign.get(a.get("account_id"), [])),
                a.get("notes") or "",
            ])
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        fname = f"RouteMail_Email_Accounts_{today}.{fmt}"
        content = _xlsx(headers, rows, "Email Accounts") if fmt == "xlsx" else _csv_bytes(headers, rows)
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if fmt == "xlsx" else "text/csv"
        return StreamingResponse(iter([content]), media_type=media,
            headers={"Content-Disposition": f'attachment; filename="{fname}"'})

    # ───────────── 2. INFRASTRUCTURE FORECASTING ─────────────
    @router.get("/forecast")
    async def forecast(monthly_target: int = Query(1_500_000, ge=0, le=100_000_000),
                       preferred_inboxes_per_domain: int = Query(5, ge=1, le=100),
                       user=Depends(get_infra_user)):
        rows = await load_inboxes_fn(db, user)
        projection = await build_projection(db, user, window_days=120)
        cap30 = aggregate_capacity(rows, projection, 30)
        cap60 = aggregate_capacity(rows, projection, 60)
        cap90 = aggregate_capacity(rows, projection, 90)

        active = [r for r in rows if r["status"] not in ("Paused", "Risky")]
        warming = [r for r in rows if r["status"] == "Warming Up"]
        domains = sorted({r["domain"] for r in rows if r["domain"]})
        active_domains = sorted({r["domain"] for r in active if r["domain"]})
        total_daily = sum(int(r["daily_limit"]) for r in active)
        monthly_capacity = total_daily * 30

        shortfall = max(monthly_target - monthly_capacity, 0)
        # Median daily limit drives the per-inbox monthly capacity.
        # The user explicitly tells us how many inboxes they want PER NEW
        # DOMAIN (via `preferred_inboxes_per_domain`) — we must NOT silently
        # fall back to the empirical current ratio (that produced unrealistic
        # numbers like "327 domains" for a 19-inbox / 6-domain pool).
        from statistics import median
        med_limit = int(median([r["daily_limit"] for r in active])) if active else 50
        per_inbox_monthly = max(med_limit * 30, 1)
        per_domain_monthly = per_inbox_monthly * max(int(preferred_inboxes_per_domain), 1)
        additional_inboxes = -(-shortfall // per_inbox_monthly) if shortfall else 0
        additional_domains = -(-shortfall // per_domain_monthly) if shortfall else 0
        projected_capacity_after = monthly_capacity + additional_inboxes * per_inbox_monthly

        return {
            "summary": {
                "total_domains": len(domains),
                "total_inboxes": len(rows),
                "active_domains": len(active_domains),
                "active_inboxes": len(active),
                "warming_inboxes": len(warming),
                "total_daily_capacity": total_daily,
                "total_monthly_capacity": monthly_capacity,
            },
            "capacity": {
                "next_30_days": cap30["month_30"],
                "next_60_days": cap60["window"],
                "next_90_days": cap90["window"],
            },
            "gap": {
                "target_monthly": monthly_target,
                "current_monthly": monthly_capacity,
                "shortfall_monthly": shortfall,
            },
            "recommendation": {
                "additional_inboxes": int(additional_inboxes),
                "additional_domains": int(additional_domains),
                "median_daily_limit": med_limit,
                "preferred_inboxes_per_domain": int(preferred_inboxes_per_domain),
                "monthly_capacity_per_inbox": int(per_inbox_monthly),
                "monthly_capacity_per_domain": int(per_domain_monthly),
                "estimated_capacity_after_expansion": int(projected_capacity_after),
            },
        }

    # ───────────── 6. DOMAIN TRACKING (CRUD) ─────────────
    def _domain_age_days(d: Dict[str, Any]) -> Optional[int]:
        added = d.get("date_added")
        if not added: return None
        try: return (datetime.now(timezone.utc).date() - _date_cls.fromisoformat(added[:10])).days
        except Exception: return None

    def _days_to_expiry(d: Dict[str, Any]) -> Optional[int]:
        exp = d.get("expiry_date")
        if not exp: return None
        try: return (_date_cls.fromisoformat(exp[:10]) - datetime.now(timezone.utc).date()).days
        except Exception: return None

    def _expiry_bucket(days: Optional[int]) -> str:
        if days is None: return "unknown"
        if days < 0: return "expired"
        if days <= 7: return "expires_7_days"
        if days <= 30: return "expires_30_days"
        if days <= 60: return "expires_60_days"
        if days <= 90: return "expires_90_days"
        return "healthy"

    @router.get("/domains")
    async def list_domains(user=Depends(get_infra_user)):
        is_admin = user.get("role") == "super_admin"
        q = {} if is_admin else {"user_id": user["user_id"]}
        docs = await db.tracked_domains.find(q, {"_id": 0}).sort("domain", 1).to_list(10000)
        enriched = []
        for d in docs:
            days = _days_to_expiry(d)
            enriched.append({
                **d,
                "days_in_infrastructure": _domain_age_days(d),
                "days_to_expiry": days,
                "expiry_bucket": _expiry_bucket(days),
            })
        # Dashboard buckets
        cnt = {"total": len(enriched), "active": 0, "expiring_90": 0, "expiring_60": 0,
               "expiring_30": 0, "expiring_7": 0, "expired": 0}
        for r in enriched:
            b = r["expiry_bucket"]
            if b == "expired": cnt["expired"] += 1
            elif b == "expires_7_days": cnt["expiring_7"] += 1; cnt["active"] += 1
            elif b == "expires_30_days": cnt["expiring_30"] += 1; cnt["active"] += 1
            elif b == "expires_60_days": cnt["expiring_60"] += 1; cnt["active"] += 1
            elif b == "expires_90_days": cnt["expiring_90"] += 1; cnt["active"] += 1
            else: cnt["active"] += 1
        return {"domains": enriched, "counts": cnt}

    @router.post("/domains")
    async def upsert_domain(payload: DomainUpsert, user=Depends(get_infra_user)):
        existing = await db.tracked_domains.find_one({"user_id": user["user_id"], "domain": payload.domain})
        doc = {
            "domain": payload.domain.lower().strip(),
            "registrar": payload.registrar,
            "purchase_date": payload.purchase_date,
            "expiry_date": payload.expiry_date,
            "renewal_date": payload.renewal_date,
            "date_added": payload.date_added or datetime.now(timezone.utc).date().isoformat(),
            "notes": payload.notes,
            "user_id": user["user_id"],
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        if existing:
            await db.tracked_domains.update_one({"_id": existing["_id"]}, {"$set": doc})
            return {"message": "Updated", "domain": doc["domain"]}
        doc["domain_id"] = f"dom_{uuid.uuid4().hex[:12]}"
        doc["created_at"] = doc["updated_at"]
        await db.tracked_domains.insert_one(doc)
        return {"message": "Created", "domain": doc["domain"], "domain_id": doc["domain_id"]}

    @router.delete("/domains/{domain}")
    async def delete_domain(domain: str, user=Depends(get_infra_user)):
        res = await db.tracked_domains.delete_one({"user_id": user["user_id"], "domain": domain.lower().strip()})
        if not res.deleted_count: raise HTTPException(status_code=404, detail="Domain not tracked")
        return {"message": "Deleted"}

    # ───────────── 7. RENEWAL REPORT EXPORT ─────────────
    @router.get("/domains/renewal-report")
    async def renewal_report(format: str = Query("xlsx"), user=Depends(get_infra_user)):
        fmt = (format or "xlsx").lower()
        if fmt not in ("xlsx", "csv"):
            raise HTTPException(status_code=400, detail="format must be xlsx|csv")
        is_admin = user.get("role") == "super_admin"
        q = {} if is_admin else {"user_id": user["user_id"]}
        docs = await db.tracked_domains.find(q, {"_id": 0}).sort("domain", 1).to_list(10000)
        headers = ["Domain", "Registrar", "Expiry Date", "Days Remaining", "Status", "Renewal Date", "Notes"]
        rows = []
        for d in docs:
            days = _days_to_expiry(d)
            status = "Expired" if (days is not None and days < 0) else (
                f"Expires in {days} days" if days is not None else "Unknown"
            )
            rows.append([d.get("domain", ""), d.get("registrar") or "", d.get("expiry_date") or "",
                         days if days is not None else "", status, d.get("renewal_date") or "", d.get("notes") or ""])
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        fname = f"RouteMail_Domain_Renewal_Report_{today}.{fmt}"
        content = _xlsx(headers, rows, "Renewal Report", fill="B45309") if fmt == "xlsx" else _csv_bytes(headers, rows)
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if fmt == "xlsx" else "text/csv"
        return StreamingResponse(iter([content]), media_type=media,
            headers={"Content-Disposition": f'attachment; filename="{fname}"'})
