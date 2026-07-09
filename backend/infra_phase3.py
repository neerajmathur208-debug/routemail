"""Auto-Allocation + Capacity Planner — Phase 3 of the Infrastructure module.

Both endpoints reuse the data the Phase-2 projection engine already computes
(`_load_inboxes` + `build_projection`) so the numbers users see in the
recommendation match the dashboard exactly.

Two endpoints:
    POST /api/infrastructure/allocate  — diversification-aware inbox picker
    POST /api/infrastructure/planner   — leads/steps/duration → required inboxes
"""

from collections import defaultdict
from datetime import datetime, timedelta, timezone, date as _date_cls
from math import ceil
from statistics import median
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field


# ---------- request schemas -------------------------------------------------

class CampaignPlanRequest(BaseModel):
    """Input for the new Campaign Capacity Planner (spec items 1-16).

    Given a campaign schedule + recipients, the planner:
      1. computes the required daily send volume,
      2. auto-derives the recommended number of inboxes (or honours a manual
         override),
      3. builds a per-execution-date capacity plan across the whole pool,
      4. deducts capacity already reserved by other running/scheduled
         campaigns and drips (via the projection map),
      5. returns a full plan with contributions, spare capacity and reasons.
    """
    recipients: int = Field(..., ge=1, le=10_000_000)
    daily_send_target: Optional[int] = Field(
        None, ge=1, le=1_000_000,
        description="Daily volume; if omitted we compute recipients / execution days.",
    )
    start_date: str = Field(..., description="Local ISO date the campaign starts.")
    steps: int = Field(..., ge=1, le=20)
    delay_days_between_steps: int = Field(..., ge=0, le=365)
    sending_days: Optional[List[int]] = Field(None, description="Weekday ints (0=Mon..6=Sun).")
    domain_reserve: int = Field(10, ge=0, le=10000,
        description="Minimum campaign sending capacity to keep unused on each domain.")
    override_required_inboxes: Optional[int] = Field(
        None, ge=1, le=10000,
        description="Manual override for auto-calculated inbox count.",
    )
    min_remaining_per_inbox: int = Field(1, ge=0, le=10000)


class AllocateRequest(BaseModel):
    required: int = Field(..., ge=1, le=10000, description="How many inboxes the campaign needs")
    ownership: Optional[str] = None
    min_remaining_per_inbox: int = Field(10, ge=0, le=10000,
        description="Skip inboxes with less remaining capacity TODAY than this floor")
    domain_capacity_floor: int = Field(10, ge=0, le=10000,
        description="Skip a whole domain when its remaining capacity today drops below this")

    # ── Schedule-aware inputs (Item 5) ────────────────────────────────────
    # When set, the allocator picks inboxes with sufficient *projected*
    # capacity on the specific future days the campaign will actually send.
    # start_date + steps + delay_days_between_steps + sending_days generate
    # the exact list of execution dates.
    start_date: Optional[str] = Field(
        None, description="Local ISO date (YYYY-MM-DD) when the campaign starts."
    )
    steps: Optional[int] = Field(None, ge=1, le=20,
        description="Number of drip steps.")
    delay_days_between_steps: Optional[int] = Field(None, ge=0, le=365,
        description="Delay between consecutive steps in calendar days.")
    sending_days: Optional[List[int]] = Field(
        None, description="Weekday integers (0=Mon..6=Sun) that campaigns may run on.",
    )
    per_day_send_estimate: Optional[int] = Field(None, ge=1, le=1_000_000,
        description="Expected sends per execution day (used to test each future day's inbox has enough headroom).")


class PlannerRequest(BaseModel):
    leads: int = Field(..., ge=1, le=10_000_000)
    steps: int = Field(..., ge=1, le=20)
    duration_days: int = Field(..., ge=1, le=365)
    sending_days_per_week: int = Field(5, ge=1, le=7)
    # Phase 3 Batch 2 — new inputs
    daily_limit_per_inbox: Optional[int] = Field(None, ge=1, le=10_000,
        description="Override the median daily_limit (e.g. 50 emails/day/inbox).")
    preferred_inboxes_per_domain: Optional[int] = Field(None, ge=1, le=100,
        description="Diversification target — how many inboxes you want to keep on each domain.")


# -- Batch-based weekly sending planner --
class BatchPlannerRequest(BaseModel):
    leads: int = Field(..., ge=1, le=10_000_000)
    steps: int = Field(..., ge=1, le=20)
    delay_days: int = Field(7, ge=1, le=90,
        description="Days between consecutive steps (uniform per spec). 7 = same weekday next week.")
    sending_days: List[int] = Field(
        default_factory=lambda: [0, 1, 2, 3, 4],
        description="Weekday integers (0=Mon..6=Sun) the user wants to send on.",
    )
    start_date: str = Field(..., description="Local start date for batch 1, YYYY-MM-DD.")
    timezone_name: str = Field("UTC", description="IANA timezone label (used for export labels — calendar math is local-date only).")

    # ONE OF the two inbox-pool inputs must be supplied. If `account_ids` is
    # set we draw the daily limit + capacity from the real Infrastructure
    # data; otherwise we use the manual `accounts` × `daily_limit_per_account`.
    account_ids: Optional[List[str]] = None
    accounts: Optional[int] = Field(None, ge=1, le=10_000)
    daily_limit_per_account: Optional[int] = Field(None, ge=1, le=10_000)


# ---------- helpers --------------------------------------------------------

def _allocate(
    inboxes: List[Dict[str, Any]],
    projection: Dict[str, Dict[str, int]],
    required: int,
    min_remaining_per_inbox: int,
    domain_capacity_floor: int,
    execution_dates: Optional[List[str]] = None,
    per_day_send_estimate: Optional[int] = None,
) -> Dict[str, Any]:
    """Diversification-aware allocator.

    Priority order:
      1. Use **one inbox per domain** before reusing any domain.
      2. Among same-priority inboxes, prefer the one with the highest
         `remaining_capacity` today (= least loaded).
      3. Skip domains whose aggregate remaining capacity today is below
         `domain_capacity_floor` (protects domains close to exhaustion).
      4. Skip inboxes whose individual `remaining_capacity` is below
         `min_remaining_per_inbox` or whose status is Paused / Risky /
         Fully Reserved.

    Warmup is NOT a blocker. An inbox that is currently warming up can still
    be allocated to a campaign — warmup and campaign sending run as two
    independent background processes and share the same ``daily_limit``
    budget. The only signals that block allocation are hard blockers:
    disconnected/errored (Risky), user-paused (Paused), or exhausted
    (Fully Reserved).

    Schedule-aware mode:
      When ``execution_dates`` is provided (as YYYY-MM-DD strings drawn from
      the campaign's start_date + delay_days_between_steps + sending_days),
      ``per_day_send_estimate`` represents the **group total** the campaign
      must be able to send **on each execution day** — NOT the required
      capacity per inbox. We divide it evenly across the requested inbox
      count to derive a per-inbox floor
      (``ceil(per_day_send_estimate / required)``); then we simulate the
      pooled distribution across the chosen inboxes to confirm the group's
      combined projected capacity meets the target on every execution date.
      Inboxes contributing only partial capacity are still eligible.
    """
    SKIP_STATUSES = {"Paused", "Risky", "Fully Reserved"}
    # Group-target semantics: per_day_send_estimate is TOTAL required per
    # execution day across the whole allocated group. Split it evenly across
    # the requested inbox count to get the per-inbox floor.
    from math import ceil as _ceil
    est_total = int(per_day_send_estimate or 0)
    per_day_floor = (
        max(1, _ceil(est_total / max(required, 1))) if est_total else min_remaining_per_inbox
    )

    def _has_future_capacity(r: Dict[str, Any]) -> bool:
        if not execution_dates:
            return True
        proj = projection.get(r["account_id"]) or {}
        for d in execution_dates:
            if int(proj.get(d, 0)) < per_day_floor:
                return False
        return True

    # Group inboxes by domain — and track exclusions with reasons for the
    # UI debug pane.
    by_domain: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    schedule_filtered_out = 0
    excluded: List[Dict[str, Any]] = []
    for r in inboxes:
        if r["status"] in SKIP_STATUSES:
            excluded.append({
                "account_id": r["account_id"],
                "email": r.get("email"),
                "domain": r.get("domain"),
                "reason": r["status"].lower(),   # paused / risky / fully reserved
            })
            continue
        if r["remaining_capacity"] < min_remaining_per_inbox:
            excluded.append({
                "account_id": r["account_id"],
                "email": r.get("email"),
                "domain": r.get("domain"),
                "reason": "no remaining campaign capacity",
            })
            continue
        if not r.get("domain"):
            excluded.append({
                "account_id": r["account_id"],
                "email": r.get("email"),
                "domain": r.get("domain"),
                "reason": "missing domain",
            })
            continue
        if not _has_future_capacity(r):
            schedule_filtered_out += 1
            excluded.append({
                "account_id": r["account_id"],
                "email": r.get("email"),
                "domain": r.get("domain"),
                "reason": f"projected capacity < {per_day_floor}/day on one or more execution dates",
            })
            continue
        by_domain[r["domain"]].append(r)

    # Drop domains whose aggregate today-remaining is below the floor.
    eligible_domains: Dict[str, List[Dict[str, Any]]] = {}
    skipped_for_capacity: List[str] = []
    for dom, rows in by_domain.items():
        dom_remaining = sum(x["remaining_capacity"] for x in rows)
        if dom_remaining < domain_capacity_floor:
            skipped_for_capacity.append(dom)
            continue
        # Sort within each domain by remaining capacity desc (then lowest
        # projected_window_total to bias toward least-loaded long-term).
        rows.sort(
            key=lambda x: (
                -x["remaining_capacity"],
                int((projection.get(x["account_id"]) or {}).get("__total__", 0)) or
                sum((projection.get(x["account_id"]) or {}).values()),
            )
        )
        eligible_domains[dom] = rows

    # Round-robin across domains
    picked: List[Dict[str, Any]] = []
    cursors: Dict[str, int] = {d: 0 for d in eligible_domains}
    # Sort domains so that those with the most depth (largest pool) come first
    # — this slightly improves coverage when the user asks for more inboxes
    # than there are domains, by ensuring we always have a fallback inbox.
    domain_order = sorted(eligible_domains.keys(), key=lambda d: -len(eligible_domains[d]))

    while len(picked) < required and any(
        cursors[d] < len(eligible_domains[d]) for d in domain_order
    ):
        progress = False
        for d in domain_order:
            if len(picked) >= required:
                break
            if cursors[d] < len(eligible_domains[d]):
                picked.append(eligible_domains[d][cursors[d]])
                cursors[d] += 1
                progress = True
        if not progress:
            break  # safety — no domain advanced this pass

    domains_used = sorted({r["domain"] for r in picked})
    avg = round(len(picked) / len(domains_used), 2) if domains_used else 0.0
    eligible_total = sum(len(v) for v in eligible_domains.values())

    warnings: List[str] = []
    if len(picked) < required:
        warnings.append(
            f"Insufficient eligible capacity — only {len(picked)} inboxes allocatable "
            f"out of {required} requested ({eligible_total} eligible after filters)."
        )
    if len(domains_used) > 0 and len(picked) / max(len(domains_used), 1) >= 5:
        warnings.append(
            f"Low domain diversification — {len(picked)} inboxes spread across only "
            f"{len(domains_used)} domains (~{avg} per domain). Spread risk concentrates."
        )
    if skipped_for_capacity:
        warnings.append(
            f"Skipped {len(skipped_for_capacity)} domain(s) near exhaustion today: "
            + ", ".join(sorted(skipped_for_capacity)[:5])
            + ("…" if len(skipped_for_capacity) > 5 else "")
        )
    if schedule_filtered_out:
        warnings.append(
            f"Schedule-aware filter removed {schedule_filtered_out} inbox(es) "
            f"with insufficient projected capacity on the campaign's execution days "
            f"(target ≥ {per_day_floor}/day per inbox = ⌈{est_total}/{required}⌉)."
        )

    # ── Per-execution-date group-capacity breakdown ───────────────────────
    # For every future date, simulate proportional distribution of the
    # ``est_total`` (group target) across the picked inboxes based on each
    # inbox's projected availability. This is the "debug output" the user
    # asked for — surfaces exactly which inboxes contribute how much on
    # which day, plus per-domain reservation totals.
    date_breakdown: List[Dict[str, Any]] = []
    if execution_dates and picked:
        for step_idx, d in enumerate(execution_dates):
            per_inbox_avail: List[Dict[str, Any]] = []
            for p in picked:
                proj_map = projection.get(p["account_id"]) or {}
                avail = int(proj_map.get(d, 0))
                per_inbox_avail.append({
                    "account_id": p["account_id"],
                    "email": p.get("email"),
                    "domain": p.get("domain"),
                    "available": avail,
                })
            total_avail = sum(x["available"] for x in per_inbox_avail)
            target = est_total or total_avail  # if no explicit estimate, fill fully
            remaining = min(target, total_avail)
            # Proportional allocation — largest-available first, cap each at
            # min(inbox_avail, ceil(target/N)); loop to soak up any leftover.
            per_inbox_avail.sort(key=lambda x: -x["available"])
            cap_ceiling = _ceil(target / max(len(per_inbox_avail), 1))
            contributions: Dict[str, int] = {x["account_id"]: 0 for x in per_inbox_avail}
            # First pass — fair share
            for x in per_inbox_avail:
                take = min(x["available"], cap_ceiling, remaining)
                contributions[x["account_id"]] = take
                remaining -= take
                if remaining <= 0:
                    break
            # Second pass — greedy fill of leftover across inboxes with
            # unused headroom.
            i = 0
            while remaining > 0 and i < len(per_inbox_avail) * 3:
                slot = per_inbox_avail[i % len(per_inbox_avail)]
                headroom = slot["available"] - contributions[slot["account_id"]]
                if headroom > 0:
                    take = min(headroom, remaining)
                    contributions[slot["account_id"]] += take
                    remaining -= take
                i += 1

            per_domain_reserved: Dict[str, int] = defaultdict(int)
            selected_lines: List[Dict[str, Any]] = []
            for x in per_inbox_avail:
                c = contributions[x["account_id"]]
                per_domain_reserved[x["domain"]] += c
                selected_lines.append({
                    "account_id": x["account_id"],
                    "email": x["email"],
                    "domain": x["domain"],
                    "available": x["available"],
                    "contributes": c,
                })
            selected_total = sum(contributions.values())
            date_breakdown.append({
                "step_number": step_idx + 1,
                "date": d,
                "required": est_total or 0,
                "available": total_avail,
                "selected": selected_total,
                "shortfall": max(0, (est_total or 0) - selected_total),
                "inboxes": selected_lines,
                "domains_reserved": [
                    {"domain": k, "reserved": v}
                    for k, v in sorted(per_domain_reserved.items(), key=lambda kv: -kv[1])
                ],
            })
        # Surface a workspace-level warning if ANY execution day couldn't be
        # fully covered by the picked group.
        shortfall_days = [b for b in date_breakdown if b["shortfall"] > 0]
        if shortfall_days and est_total > 0:
            worst = max(shortfall_days, key=lambda b: b["shortfall"])
            warnings.append(
                f"Group capacity insufficient on {len(shortfall_days)} of "
                f"{len(execution_dates)} execution date(s). Worst: "
                f"{worst['date']} — need {worst['required']}, "
                f"selected group can cover {worst['selected']}."
            )

    return {
        "requested": required,
        "allocated": len(picked),
        "eligible_count": eligible_total,
        "inboxes": [
            {
                "account_id": p["account_id"],
                "email": p["email"],
                "domain": p["domain"],
                "ownership": p["ownership"],
                "daily_limit": p["daily_limit"],
                "remaining_capacity": p["remaining_capacity"],
                "status": p["status"],
            }
            for p in picked
        ],
        "domains_used": domains_used,
        "avg_inboxes_per_domain": avg,
        "warnings": warnings,
        "skipped_domains_near_exhaustion": skipped_for_capacity,
        # Schedule-aware breakdowns (empty lists when not in schedule mode).
        "date_breakdown": date_breakdown,
        "excluded": excluded,
        "per_inbox_floor": per_day_floor if execution_dates else None,
        "group_target_per_day": est_total or None,
    }


def _next_sending_day(d: _date_cls, sending_days: List[int]) -> _date_cls:
    """Roll `d` forward to the next allowed weekday. If `d` itself is allowed
    we return it unchanged. Caps at 14 iterations as a safety net against an
    empty sending_days list."""
    if not sending_days:
        return d
    s = set(sending_days)
    for _ in range(14):
        if d.weekday() in s:
            return d
        d = d + timedelta(days=1)
    return d


def _batch_plan(
    req: BatchPlannerRequest,
    inboxes_for_pool: List[Dict[str, Any]],
    projection: Dict[str, Dict[str, int]],
) -> Dict[str, Any]:
    """Build a per-day, per-step calendar for a batch-based campaign.

    Strategy:
      1. Pick the daily-capacity figure — sum of `daily_limit` across the
         picked account pool when `account_ids` was supplied, else the
         manual `accounts * daily_limit_per_account`.
      2. Carve `leads` into `ceil(leads / daily_capacity)` batches; the last
         batch carries any remainder.
      3. Assign batches to the next N available sending-days starting at
         `start_date` (rolled forward to the nearest allowed weekday).
      4. For each (batch, step≥2), date = batch_step1_date + delay_days*(step-1),
         rolled forward to the next allowed sending-day if it lands on a
         non-sending weekday (spec §6).
      5. For every (date, sends_required) pair, check against the real
         projection-aware remaining capacity for the pool and tag the row
         Ready / Partial Capacity / Insufficient Capacity (spec §7).
    """
    # ---- 1. Resolve the pool + daily capacity --------------------------
    if req.account_ids:
        pool = [r for r in inboxes_for_pool if r["account_id"] in set(req.account_ids)]
        if not pool:
            raise HTTPException(status_code=400, detail="None of the supplied account_ids are visible to this user.")
        pool_account_ids = [r["account_id"] for r in pool]
        daily_capacity = sum(int(r["daily_limit"]) for r in pool)
        account_count = len(pool)
        median_limit = max(int(median([r["daily_limit"] for r in pool])), 1)
        pool_label = "real_inbox_pool"
    else:
        if not (req.accounts and req.daily_limit_per_account):
            raise HTTPException(
                status_code=400,
                detail="Provide either `account_ids` OR (`accounts` AND `daily_limit_per_account`).",
            )
        pool_account_ids = []
        daily_capacity = int(req.accounts) * int(req.daily_limit_per_account)
        account_count = int(req.accounts)
        median_limit = int(req.daily_limit_per_account)
        pool_label = "manual"
        pool = []

    if daily_capacity <= 0:
        raise HTTPException(status_code=400, detail="Daily capacity must be > 0")

    # ---- 2. Carve leads into batches -----------------------------------
    total_batches = ceil(req.leads / daily_capacity)
    batches: List[Dict[str, Any]] = []
    remaining = req.leads
    for i in range(total_batches):
        size = min(daily_capacity, remaining)
        batches.append({"batch": i + 1, "leads": size})
        remaining -= size

    # ---- 3. Assign step-1 dates ----------------------------------------
    try:
        cursor = _date_cls.fromisoformat(req.start_date)
    except Exception:
        raise HTTPException(status_code=400, detail="start_date must be YYYY-MM-DD")
    sending_days = sorted({int(d) for d in (req.sending_days or [0, 1, 2, 3, 4])})

    for b in batches:
        cursor = _next_sending_day(cursor, sending_days)
        b["step_1_date"] = cursor
        cursor = cursor + timedelta(days=1)  # advance for next batch

    # ---- 4. Project every (batch, step) onto a date --------------------
    # Pre-compute the projection load for the pool — only relevant when we
    # have a real account pool. For the manual mode we assume zero existing
    # load (the user is forecasting in isolation).
    def _pool_used_on(d: _date_cls) -> int:
        if not pool_account_ids:
            return 0
        iso = d.isoformat()
        return sum(int((projection.get(aid) or {}).get(iso, 0)) for aid in pool_account_ids)

    def _pool_capacity_on(d: _date_cls) -> int:
        # For the manual pool we assume daily_capacity every day; for the real
        # pool we sum the per-inbox daily_limit (constant) and subtract the
        # already-projected load.
        if not pool_account_ids:
            return daily_capacity
        return daily_capacity - _pool_used_on(d)

    schedule: List[Dict[str, Any]] = []
    # Aggregate self-conflict — when two batches' steps land on the same day,
    # we add their loads together against the same capacity.
    self_load: Dict[str, int] = defaultdict(int)
    plan_steps: List[Dict[str, Any]] = []

    for b in batches:
        step_date = b["step_1_date"]
        for step in range(1, req.steps + 1):
            if step > 1:
                step_date = _next_sending_day(
                    step_date + timedelta(days=req.delay_days), sending_days
                )
            plan_steps.append({
                "batch": b["batch"],
                "step": step,
                "leads": b["leads"],
                "date": step_date,
            })
            self_load[step_date.isoformat()] += b["leads"]

    # Resolve status per row
    for row in plan_steps:
        iso = row["date"].isoformat()
        required = int(self_load[iso])  # combined load across overlapping batches on this date
        if pool_account_ids:
            available = max(_pool_capacity_on(row["date"]), 0)
        else:
            available = daily_capacity
        if available >= required:
            # If this row alone fits but the combined day still uses more than
            # half of capacity, surface that distinction — keep "Ready" for
            # the row level since the row's own leads fit comfortably, only
            # downgrade if the day-level overflow puts THIS row past capacity.
            status = "Ready"
        elif available >= row["leads"]:
            status = "Partial Capacity"
        else:
            status = "Insufficient Capacity"
        schedule.append({
            "date": iso,
            "weekday": row["date"].weekday(),
            "weekday_name": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"][row["date"].weekday()],
            "batch": row["batch"],
            "step": row["step"],
            "leads": row["leads"],
            "required_capacity": required,
            "available_capacity": available,
            "shortfall": max(required - available, 0),
            "status": status,
        })
    schedule.sort(key=lambda r: (r["date"], r["batch"], r["step"]))

    # ---- 5. Roll-ups ---------------------------------------------------
    total_emails = req.leads * req.steps
    last_date = schedule[-1]["date"] if schedule else req.start_date
    first_date = schedule[0]["date"] if schedule else req.start_date
    duration_days = (
        _date_cls.fromisoformat(last_date) - _date_cls.fromisoformat(first_date)
    ).days + 1

    warnings: List[str] = []
    bad_days = [r for r in schedule if r["status"] != "Ready"]
    if bad_days:
        for r in bad_days[:5]:
            warnings.append(
                f"Capacity exceeded on {r['date']}. "
                f"Required: {r['required_capacity']:,} · Available: {r['available_capacity']:,} · "
                f"Shortfall: {r['shortfall']:,}"
            )
        if len(bad_days) > 5:
            warnings.append(f"…and {len(bad_days) - 5} more day(s) with capacity issues.")

    overall_status = "Ready"
    if any(r["status"] == "Insufficient Capacity" for r in schedule):
        overall_status = "Insufficient Capacity"
    elif any(r["status"] == "Partial Capacity" for r in schedule):
        overall_status = "Partial Capacity"

    return {
        "inputs": {
            "leads": req.leads,
            "steps": req.steps,
            "delay_days": req.delay_days,
            "sending_days": sending_days,
            "start_date": req.start_date,
            "timezone": req.timezone_name,
            "pool_source": pool_label,
            "pool_account_ids": pool_account_ids,
            "manual_accounts": req.accounts,
            "manual_daily_limit_per_account": req.daily_limit_per_account,
        },
        "summary": {
            "total_leads": req.leads,
            "total_batches": total_batches,
            "total_steps": req.steps,
            "total_emails": total_emails,
            "daily_capacity": daily_capacity,
            "account_count": account_count,
            "median_daily_limit": median_limit,
            "duration_days": duration_days,
            "first_send_date": first_date,
            "last_send_date": last_date,
            "status": overall_status,
        },
        "batches": [
            {
                "batch": b["batch"],
                "leads": b["leads"],
                "step_1_date": b["step_1_date"].isoformat(),
                "weekday_name": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"][b["step_1_date"].weekday()],
            }
            for b in batches
        ],
        "schedule": schedule,
        "warnings": warnings,
    }



def _plan(
    inboxes: List[Dict[str, Any]],
    capacity: Dict[str, int],
    req: PlannerRequest,
) -> Dict[str, Any]:
    """Capacity Planner math.

    Returns enough numbers for the UI to clearly say either "Ready" or
    "Insufficient Capacity — need N more inboxes" along with a few extra
    diagnostic warnings.
    """
    total_emails = req.leads * req.steps
    sending_days_in_window = int(round(req.duration_days * (req.sending_days_per_week / 7.0)))
    sending_days_in_window = max(sending_days_in_window, 1)
    required_daily_volume = ceil(total_emails / sending_days_in_window)

    # Use the **median** daily_limit across eligible (non-paused, non-risky)
    # inboxes — robust to a couple of outlier inboxes with very high limits.
    # Warmup is deliberately NOT a blocker here: a warming inbox still
    # contributes real send capacity for campaign planning.
    eligible = [
        r for r in inboxes
        if r["status"] not in {"Paused", "Risky"}
    ]
    if eligible:
        median_limit = max(int(median(r["daily_limit"] for r in eligible)), 1)
    else:
        median_limit = 50  # sane fallback if there are no usable inboxes at all

    # Phase 3 Batch 2 — explicit override beats the empirical median.
    med_limit = int(req.daily_limit_per_inbox or median_limit)

    required_inboxes = ceil(required_daily_volume / med_limit)
    available_inboxes = len(eligible)
    additional_needed = max(0, required_inboxes - available_inboxes)

    # Phase 3 Batch 2 — domain math driven by the diversification target.
    preferred_inboxes_per_domain = int(req.preferred_inboxes_per_domain or 5)
    required_domains = max(1, ceil(required_inboxes / preferred_inboxes_per_domain))
    daily_capacity_per_domain = med_limit * preferred_inboxes_per_domain
    daily_capacity_total = med_limit * required_inboxes

    # Existing infrastructure snapshot for "current vs required" recommendations.
    current_inboxes = len(eligible)
    current_domains = len({r["domain"] for r in eligible if r.get("domain")})
    if current_domains == 0:
        current_avg_inboxes_per_domain = 0
    else:
        current_avg_inboxes_per_domain = round(current_inboxes / current_domains, 1)
    current_daily_per_domain = (
        round(sum(r["daily_limit"] for r in eligible) / current_domains)
        if current_domains else 0
    )
    additional_domains_needed = max(0, required_domains - current_domains)

    today = capacity.get("today", 0)
    window = capacity.get("window", capacity.get("month_30", 0))

    # Did the user buy themselves enough total capacity across the window?
    enough_window = (window or 0) >= total_emails
    enough_inboxes = required_inboxes <= available_inboxes
    ready = enough_window and enough_inboxes

    # Estimated completion = ceil(total_emails / per-day-capacity-of-allocated)
    per_day_capacity = required_inboxes * med_limit if required_inboxes > 0 else med_limit
    actual_sending_days = ceil(total_emails / max(per_day_capacity, 1))
    # Calendar days = actual_sending_days × 7 / sending_days_per_week
    estimated_completion_days = ceil(actual_sending_days * 7 / req.sending_days_per_week)

    warnings: List[str] = []
    if additional_needed > 0:
        warnings.append(
            f"Need {required_inboxes} inboxes; available {available_inboxes}. "
            f"Add {additional_needed} more to fit the schedule."
        )
    if not enough_window:
        shortfall = total_emails - (window or 0)
        warnings.append(
            f"Window capacity short by {shortfall:,} emails over {req.duration_days} days."
        )
    # Domain diversification warning — if all eligible inboxes live on ≤ 4 domains
    domain_count = len({r["domain"] for r in eligible if r.get("domain")})
    if available_inboxes > 0 and domain_count <= 4:
        warnings.append(
            f"Low domain diversification — only {domain_count} domain(s) across "
            f"{available_inboxes} eligible inboxes. Spread risk concentrates."
        )

    # Phase 3 Batch 2 — concrete diversification check against the preferred
    # target (default 5 inboxes / domain → 250 emails / domain / day).
    if current_daily_per_domain > daily_capacity_per_domain * 1.5:
        warnings.append(
            f"Current sending pace is {current_daily_per_domain}/day/domain, "
            f"target is {daily_capacity_per_domain}/day/domain. Add more domains."
        )
    if additional_domains_needed > 0:
        warnings.append(
            f"Need {required_domains} domains for the target; you have {current_domains}. "
            f"Add {additional_domains_needed} more."
        )

    return {
        "inputs": {
            "leads": req.leads,
            "steps": req.steps,
            "duration_days": req.duration_days,
            "sending_days_per_week": req.sending_days_per_week,
            "daily_limit_per_inbox": med_limit,
            "preferred_inboxes_per_domain": preferred_inboxes_per_domain,
        },
        "outputs": {
            "total_emails": total_emails,
            "sending_days_in_window": sending_days_in_window,
            "required_daily_volume": required_daily_volume,
            "required_inboxes": required_inboxes,
            "required_domains": required_domains,
            "daily_capacity_total": daily_capacity_total,
            "daily_capacity_per_domain": daily_capacity_per_domain,
            "daily_sends_per_inbox": med_limit,
            "available_inboxes": available_inboxes,
            "additional_inboxes_required": additional_needed,
            "additional_domains_required": additional_domains_needed,
            "current_inboxes": current_inboxes,
            "current_domains": current_domains,
            "current_avg_inboxes_per_domain": current_avg_inboxes_per_domain,
            "current_daily_per_domain": current_daily_per_domain,
            "median_daily_limit": median_limit,
            "available_capacity_today": today,
            "available_capacity_window": window,
            "estimated_completion_days": estimated_completion_days,
            "domain_diversity": domain_count,
        },
        "status": "Ready" if ready else "Insufficient Capacity",
        "warnings": warnings,
    }


# ---------- router builder -------------------------------------------------

def _compute_execution_dates(start_date_iso: str, steps: int, delay_days: int,
                              sending_days: Optional[List[int]]) -> List[str]:
    """Reusable: build the exact execution date list for a campaign,
    rolling non-sending weekdays forward."""
    from datetime import date, timedelta as _td
    start = date.fromisoformat(start_date_iso)
    allowed = set(sending_days or list(range(7)))
    if not allowed:
        allowed = set(range(7))
    dates: List[str] = []
    cursor = start
    for step_idx in range(steps):
        if step_idx > 0:
            cursor = cursor + _td(days=delay_days)
        safety = 0
        while cursor.weekday() not in allowed and safety < 14:
            cursor = cursor + _td(days=1)
            safety += 1
        dates.append(cursor.isoformat())
    return dates


def _plan_campaign(
    inboxes: List[Dict[str, Any]],
    projection: Dict[str, Dict[str, int]],
    execution_dates: List[str],
    daily_target: int,
    domain_reserve: int,
    override_required: Optional[int] = None,
    min_remaining_per_inbox: int = 1,
) -> Dict[str, Any]:
    """Campaign Capacity Planner.

    Distinct from ``_allocate`` (which is a same-day inbox filter): the
    planner reasons across the ENTIRE execution calendar. Its job is to pick
    a set of inboxes whose *combined* projected daily capacity on every one
    of ``execution_dates`` meets ``daily_target`` — allowing partial
    contributions from each inbox and slight over-allocation.

    Warmup is intentionally NOT a filter here. Only these exclude an inbox:
      * disconnected/errored (Risky)
      * user-paused (Paused)
      * zero projected capacity on ALL execution dates (Fully Reserved)
    """
    excluded: List[Dict[str, Any]] = []
    hard_block = {"Paused", "Risky"}

    # Filter to allocatable pool
    pool: List[Dict[str, Any]] = []
    for r in inboxes:
        if r.get("status") in hard_block:
            excluded.append({
                "account_id": r["account_id"], "email": r.get("email"),
                "domain": r.get("domain"),
                "reason": r["status"].lower(),
            })
            continue
        if not r.get("domain"):
            excluded.append({
                "account_id": r["account_id"], "email": r.get("email"),
                "domain": None, "reason": "missing domain",
            })
            continue
        proj_map = projection.get(r["account_id"]) or {}
        # Total projected capacity across the campaign's execution days
        total_across_dates = sum(int(proj_map.get(d, 0)) for d in execution_dates)
        if total_across_dates <= 0:
            excluded.append({
                "account_id": r["account_id"], "email": r.get("email"),
                "domain": r.get("domain"),
                "reason": "no remaining projected campaign capacity on execution dates",
            })
            continue
        # Note the specific dates where the inbox is at zero (informational).
        zero_dates = [d for d in execution_dates if int(proj_map.get(d, 0)) <= 0]
        if len(zero_dates) == len(execution_dates):
            # already handled above but defensive
            continue
        r["_zero_dates"] = zero_dates
        r["_avg_daily_projected"] = int(total_across_dates / max(len(execution_dates), 1))
        r["_min_daily_projected"] = min(int(proj_map.get(d, 0)) for d in execution_dates)
        pool.append(r)

    # Step 1 — auto-recommend inbox count based on median projected capacity.
    if pool:
        median_avail = int(median([r["_min_daily_projected"] for r in pool])) or 1
    else:
        median_avail = 1
    auto_recommended = max(1, ceil(daily_target / max(median_avail, 1)))
    requested_count = int(override_required or auto_recommended)

    # Step 2 — sort pool for diversification-first pick (unique domains first,
    # then by min-projected desc so the healthiest inboxes come first).
    pool.sort(key=lambda r: (r.get("_min_daily_projected", 0)), reverse=True)
    by_domain: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in pool:
        by_domain[r["domain"]].append(r)

    # Round-robin domain pick until we have `requested_count` picks or exhaust pool.
    picked: List[Dict[str, Any]] = []
    cursors: Dict[str, int] = {d: 0 for d in by_domain}
    domain_order = sorted(by_domain.keys(), key=lambda d: -len(by_domain[d]))
    while len(picked) < requested_count and any(
        cursors[d] < len(by_domain[d]) for d in domain_order
    ):
        progress = False
        for d in domain_order:
            if len(picked) >= requested_count:
                break
            if cursors[d] < len(by_domain[d]):
                picked.append(by_domain[d][cursors[d]])
                cursors[d] += 1
                progress = True
        if not progress:
            break

    # Step 3 — simulate per-execution-date proportional distribution.
    from collections import defaultdict as _dd
    per_inbox_totals: Dict[str, int] = _dd(int)
    per_inbox_by_date: Dict[str, Dict[str, int]] = _dd(dict)
    per_domain_by_date: Dict[str, Dict[str, int]] = _dd(lambda: _dd(int))
    date_plan: List[Dict[str, Any]] = []

    for step_idx, d in enumerate(execution_dates):
        per_inbox = []
        for p in picked:
            avail = int((projection.get(p["account_id"]) or {}).get(d, 0))
            per_inbox.append({
                "account_id": p["account_id"], "email": p.get("email"),
                "domain": p.get("domain"), "available": avail,
            })
        total_avail = sum(x["available"] for x in per_inbox)
        need = daily_target
        # Enforce domain reserve — for each domain, cap the *total* usable
        # capacity at (pool - reserve). Scale per-inbox availability
        # proportionally so their sum respects the domain cap.
        from collections import defaultdict as _dd2
        domain_avail: Dict[str, int] = _dd2(int)
        for x in per_inbox:
            domain_avail[x["domain"]] += x["available"]
        domain_cap: Dict[str, int] = {}
        for dom, pool_tot in domain_avail.items():
            domain_cap[dom] = max(0, pool_tot - domain_reserve)
        for dom, cap in domain_cap.items():
            if cap >= domain_avail[dom]:
                continue  # nothing to cap
            # Scale each inbox in this domain proportionally.
            pool_tot = domain_avail[dom]
            if pool_tot <= 0:
                continue
            for x in per_inbox:
                if x["domain"] == dom:
                    x["available"] = int(x["available"] * cap / pool_tot)
        total_avail_after_reserve = sum(x["available"] for x in per_inbox)
        # Proportional distribution: largest-first cap at ceil(need/N), then leftover fill.
        n = max(len(per_inbox), 1)
        cap_ceiling = ceil(need / n)
        per_inbox.sort(key=lambda x: -x["available"])
        contributions: Dict[str, int] = {x["account_id"]: 0 for x in per_inbox}
        remaining = min(need, total_avail_after_reserve)
        for x in per_inbox:
            take = min(x["available"], cap_ceiling, remaining)
            contributions[x["account_id"]] = take
            remaining -= take
            if remaining <= 0:
                break
        # Greedy leftover fill
        i = 0
        while remaining > 0 and i < n * 3:
            slot = per_inbox[i % n]
            headroom = slot["available"] - contributions[slot["account_id"]]
            if headroom > 0:
                take = min(headroom, remaining)
                contributions[slot["account_id"]] += take
                remaining -= take
            i += 1

        selected_lines = []
        per_day_domain = _dd(int)
        for x in per_inbox:
            c = contributions[x["account_id"]]
            per_inbox_totals[x["account_id"]] += c
            per_inbox_by_date[x["account_id"]][d] = c
            per_day_domain[x["domain"]] += c
            selected_lines.append({**x, "contributes": c})
        selected_total = sum(contributions.values())
        for dom, tot in per_day_domain.items():
            per_domain_by_date[d][dom] = tot
        date_plan.append({
            "step_number": step_idx + 1,
            "date": d,
            "required": need,
            "available": total_avail,
            "available_after_reserve": total_avail_after_reserve,
            "selected": selected_total,
            "shortfall": max(0, need - selected_total),
            "inboxes": selected_lines,
            "domains_reserved": [
                {"domain": dom, "reserved": tot}
                for dom, tot in sorted(per_day_domain.items(), key=lambda kv: -kv[1])
            ],
        })

    # Aggregated view per selected inbox (used by the UI).
    inbox_summaries = []
    for p in picked:
        aid = p["account_id"]
        inbox_summaries.append({
            "account_id": aid,
            "email": p.get("email"),
            "domain": p.get("domain"),
            "daily_limit": p.get("daily_limit"),
            "min_projected": p.get("_min_daily_projected", 0),
            "avg_projected": p.get("_avg_daily_projected", 0),
            "allocated_total": per_inbox_totals.get(aid, 0),
            "per_date": per_inbox_by_date.get(aid, {}),
            "dates_covered": [
                d for d in execution_dates
                if per_inbox_by_date.get(aid, {}).get(d, 0) > 0
            ],
        })

    domains_used = sorted({p["domain"] for p in picked})
    combined_min_daily = min((b["selected"] for b in date_plan), default=0)
    projected_spare = max(0, combined_min_daily - daily_target)
    total_shortfall_days = sum(1 for b in date_plan if b["shortfall"] > 0)

    warnings: List[str] = []
    if total_shortfall_days > 0:
        warnings.append(
            f"Group capacity insufficient on {total_shortfall_days} of "
            f"{len(execution_dates)} execution date(s). Consider more inboxes."
        )
    if not picked:
        warnings.append("No allocatable inboxes matched the campaign schedule.")
    if len(picked) < requested_count:
        warnings.append(
            f"Only {len(picked)} inboxes available; recommended {requested_count}."
        )

    status = "Healthy" if total_shortfall_days == 0 and picked else (
        "Partial" if picked else "Insufficient"
    )
    return {
        "recipients": None,  # populated by endpoint
        "daily_target": daily_target,
        "auto_recommended_inboxes": auto_recommended,
        "requested_inboxes": requested_count,
        "median_daily_projected": median_avail,
        "execution_dates": execution_dates,
        "recommended_inboxes_count": len(picked),
        "domains_used": domains_used,
        "combined_min_daily_capacity": combined_min_daily,
        "projected_spare_capacity": projected_spare,
        "shortfall_days": total_shortfall_days,
        "infrastructure_status": status,
        "inboxes": inbox_summaries,
        "date_plan": date_plan,
        "excluded": excluded,
        "warnings": warnings,
    }


def attach_phase3_routes(router: APIRouter, db, get_infra_user, load_inboxes_fn, build_projection_fn, aggregate_capacity_fn):
    """Bolt the Phase-3 endpoints onto the existing /infrastructure router.

    Why we pass the helpers in rather than importing them directly: the
    Phase-2 helpers live inside `infrastructure_routes.build_infrastructure_router`
    (closure access to `db`). Passing them as args avoids a circular import.
    """

    @router.post("/allocate")
    async def allocate(req: AllocateRequest, user=Depends(get_infra_user)):
        rows = await load_inboxes_fn(db, user)
        projection = await build_projection_fn(db, user, window_days=120)
        # enrich rows with projection rollup (matches /inboxes output shape)
        for r in rows:
            r["projected_window_total"] = sum((projection.get(r["account_id"]) or {}).values())

        # ── Compute execution dates from the campaign schedule ────────────
        # When the caller supplies start_date + steps + delay_days_between_steps
        # (+ optional sending_days weekday filter), we generate the exact
        # future dates the campaign will send on and pass them to the
        # allocator. If any of those dates fall on a non-sending weekday,
        # we roll forward to the next allowed weekday (matching how the
        # drip worker actually behaves).
        execution_dates: Optional[List[str]] = None
        if req.start_date and req.steps and req.delay_days_between_steps is not None:
            try:
                from datetime import date, timedelta as _td
                start = date.fromisoformat(req.start_date)
                allowed_weekdays = set(req.sending_days or list(range(7)))
                if not allowed_weekdays:
                    allowed_weekdays = set(range(7))
                dates: List[str] = []
                cursor = start
                for step_idx in range(req.steps):
                    if step_idx > 0:
                        cursor = cursor + _td(days=req.delay_days_between_steps)
                    # Roll forward to the next allowed sending day.
                    safety = 0
                    while cursor.weekday() not in allowed_weekdays and safety < 14:
                        cursor = cursor + _td(days=1)
                        safety += 1
                    dates.append(cursor.isoformat())
                execution_dates = dates
            except Exception:
                # Malformed date input — degrade gracefully to same-day only.
                execution_dates = None

        result = _allocate(
            rows, projection, req.required,
            req.min_remaining_per_inbox, req.domain_capacity_floor,
            execution_dates=execution_dates,
            per_day_send_estimate=req.per_day_send_estimate,
        )
        if execution_dates:
            result["execution_dates"] = execution_dates
        return result

    @router.post("/plan-campaign")
    async def plan_campaign(req: CampaignPlanRequest, user=Depends(get_infra_user)):
        """Full Campaign Capacity Planner (spec items 1-16).

        Redesign of Auto-Allocate as a capacity-planning engine:
          * auto-derives the recommended inbox count from median projected
            per-inbox capacity vs the daily send target,
          * simulates per-execution-date proportional distribution across
            the picked inboxes,
          * accepts partial contributions,
          * excludes only hard blockers (Paused / Risky / zero-projection),
          * ignores warmup + domain score entirely,
          * respects per-domain reserve.
        """
        rows = await load_inboxes_fn(db, user)
        projection = await build_projection_fn(db, user, window_days=120)
        execution_dates = _compute_execution_dates(
            req.start_date, req.steps, req.delay_days_between_steps,
            req.sending_days,
        )
        # Derive the effective daily target: user-supplied or recipients / #dates
        daily_target = int(req.daily_send_target or ceil(req.recipients / max(len(execution_dates), 1)))
        plan = _plan_campaign(
            rows, projection, execution_dates, daily_target,
            domain_reserve=req.domain_reserve,
            override_required=req.override_required_inboxes,
            min_remaining_per_inbox=req.min_remaining_per_inbox,
        )
        plan["recipients"] = req.recipients
        return plan

    @router.post("/planner")
    async def planner(req: PlannerRequest, user=Depends(get_infra_user)):
        rows = await load_inboxes_fn(db, user)
        projection = await build_projection_fn(db, user, window_days=120)
        capacity = aggregate_capacity_fn(rows, projection, 120)
        return _plan(rows, capacity, req)

    @router.post("/planner/export")
    async def planner_export(
        req: PlannerRequest,
        format: str = Query("xlsx", description="xlsx | csv"),
        user=Depends(get_infra_user),
    ):
        """Export the Capacity Planner outputs as XLSX or CSV (Phase 3 Batch 2)."""
        from fastapi.responses import Response as _Resp
        import csv
        import io
        rows = await load_inboxes_fn(db, user)
        projection = await build_projection_fn(db, user, window_days=120)
        capacity = aggregate_capacity_fn(rows, projection, 120)
        plan = _plan(rows, capacity, req)
        inp = plan["inputs"]
        out = plan["outputs"]
        headers = ["Field", "Value"]
        body_rows = [
            ["Leads", inp["leads"]],
            ["Steps", inp["steps"]],
            ["Duration (days)", inp["duration_days"]],
            ["Sending days / week", inp["sending_days_per_week"]],
            ["Daily limit / inbox", inp["daily_limit_per_inbox"]],
            ["Preferred inboxes / domain", inp["preferred_inboxes_per_domain"]],
            ["", ""],
            ["Required Inboxes", out["required_inboxes"]],
            ["Required Domains", out["required_domains"]],
            ["Daily Capacity (total)", out["daily_capacity_total"]],
            ["Daily Capacity / Domain", out["daily_capacity_per_domain"]],
            ["Daily Sends / Inbox", out["daily_sends_per_inbox"]],
            ["Current Inboxes", out["current_inboxes"]],
            ["Current Domains", out["current_domains"]],
            ["Additional Inboxes Needed", out["additional_inboxes_required"]],
            ["Additional Domains Needed", out["additional_domains_required"]],
            ["Estimated Completion (days)", out["estimated_completion_days"]],
            ["Status", plan["status"]],
        ]
        for w in plan.get("warnings", []):
            body_rows.append(["Warning", w])
        if format == "csv":
            buf = io.StringIO()
            wcsv = csv.writer(buf)
            wcsv.writerow(headers)
            for r in body_rows:
                wcsv.writerow(r)
            return _Resp(
                content=buf.getvalue(),
                media_type="text/csv",
                headers={"Content-Disposition": 'attachment; filename="capacity-planner.csv"'},
            )
        # xlsx
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Capacity Planner"
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(fgColor="4338CA", fill_type="solid")
        for r in body_rows:
            ws.append(r)
        out_bytes = io.BytesIO()
        wb.save(out_bytes)
        return _Resp(
            content=out_bytes.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="capacity-planner.xlsx"'},
        )

    @router.post("/planner/batch")
    async def planner_batch(req: BatchPlannerRequest, user=Depends(get_infra_user)):
        rows = await load_inboxes_fn(db, user)
        projection = await build_projection_fn(db, user, window_days=120)
        return _batch_plan(req, rows, projection)

    @router.post("/planner/batch/export")
    async def planner_batch_export(
        req: BatchPlannerRequest,
        format: str = Query("xlsx", description="xlsx | csv"),
        user=Depends(get_infra_user),
    ):
        rows = await load_inboxes_fn(db, user)
        projection = await build_projection_fn(db, user, window_days=120)
        plan = _batch_plan(req, rows, projection)
        fmt = (format or "xlsx").lower()
        if fmt not in ("xlsx", "csv"):
            raise HTTPException(status_code=400, detail="format must be xlsx|csv")

        import io
        import csv as _csv
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        fname = f"RouteMail_Batch_Plan_{today}.{fmt}"
        HEADERS = [
            "Date", "Day", "Batch", "Step", "Leads Scheduled",
            "Required Capacity", "Available Capacity", "Shortfall", "Status",
        ]

        if fmt == "xlsx":
            wb = Workbook()
            # Schedule sheet
            ws = wb.active
            ws.title = "Schedule"
            font = Font(bold=True, color="FFFFFF")
            fill = PatternFill("solid", fgColor="0F766E")  # teal-700
            for i, h in enumerate(HEADERS, 1):
                c = ws.cell(row=1, column=i, value=h)
                c.font = font
                c.fill = fill
            for idx, r in enumerate(plan["schedule"], 2):
                ws.cell(row=idx, column=1, value=r["date"])
                ws.cell(row=idx, column=2, value=r["weekday_name"])
                ws.cell(row=idx, column=3, value=r["batch"])
                ws.cell(row=idx, column=4, value=r["step"])
                ws.cell(row=idx, column=5, value=r["leads"])
                ws.cell(row=idx, column=6, value=r["required_capacity"])
                ws.cell(row=idx, column=7, value=r["available_capacity"])
                ws.cell(row=idx, column=8, value=r["shortfall"])
                ws.cell(row=idx, column=9, value=r["status"])
            for col_cells in ws.columns:
                length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 28)
            # Summary sheet (first to greet the user when they open the file)
            sm = wb.create_sheet("Summary", 0)
            sm["A1"] = "RouteMail — Batch Plan"
            sm["A1"].font = Font(bold=True, size=14)
            sumrows = [
                ("Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
                ("Status", plan["summary"]["status"]),
                ("Total Leads", plan["summary"]["total_leads"]),
                ("Total Batches", plan["summary"]["total_batches"]),
                ("Steps", plan["summary"]["total_steps"]),
                ("Total Emails", plan["summary"]["total_emails"]),
                ("Daily Capacity", plan["summary"]["daily_capacity"]),
                ("Account Count", plan["summary"]["account_count"]),
                ("Median Daily Limit", plan["summary"]["median_daily_limit"]),
                ("Duration (days)", plan["summary"]["duration_days"]),
                ("First Send", plan["summary"]["first_send_date"]),
                ("Last Send", plan["summary"]["last_send_date"]),
            ]
            for r_idx, (k, v) in enumerate(sumrows, 3):
                sm.cell(row=r_idx, column=1, value=k).font = Font(bold=True)
                sm.cell(row=r_idx, column=2, value=v)
            if plan["warnings"]:
                sm.cell(row=len(sumrows) + 5, column=1, value="Warnings").font = Font(bold=True)
                for i, w in enumerate(plan["warnings"]):
                    sm.cell(row=len(sumrows) + 6 + i, column=1, value=w)
            sm.column_dimensions["A"].width = 28
            sm.column_dimensions["B"].width = 24
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            content = buf.read()
            media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            out = io.StringIO()
            w = _csv.writer(out)
            w.writerow(HEADERS)
            for r in plan["schedule"]:
                w.writerow([
                    r["date"], r["weekday_name"], r["batch"], r["step"], r["leads"],
                    r["required_capacity"], r["available_capacity"], r["shortfall"], r["status"],
                ])
            content = out.getvalue().encode()
            media = "text/csv"
        return StreamingResponse(
            iter([content]),
            media_type=media,
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
